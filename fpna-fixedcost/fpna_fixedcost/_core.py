"""
fixed_cost_decision_system.py
계약 중심 고정비 의사결정 시스템 — 단일 파일 통합본 (순수 stdlib)

포함:
  SECTION 1  공통 토대 (provenance·NPV/DCF·가중통계·그라운딩·중요성)
  SECTION 2  엔진 2.1  건물 임차 시장임대료 추정 + 유불리   [RICS 비교법]
  SECTION 3  엔진 2.2  트럭 구매 vs 리스 (세후 DCF)         [IFRS 16: IBR≠허들]
  SECTION 4  엔진 2.3  IAS 36 손상 (VIU+CGU)                [IAS 36]
  SECTION 5  투영 엔진 §7.1 (상각 roll-forward·리스 step-up·run-rate)
  SECTION 6  의사결정 → 투영 배선 (결정이 fcst 라인 스케줄로)
  SECTION 7  SOX/ICFR 통제 매트릭스 §7.13 (증적 자동수집)
  SECTION 8  Decision Analysis 카드 + 게이트 + SQLite
  SECTION 10 승인 워크플로 §6.8 (직무분리 SoD + 실행 → ICFR effective 완성)
  SECTION 11 openpyxl 보고서 (View Contract: 전체 시간축·recon·evidence health)
  SECTION 13 트리거/큐 층 (§4 work_item·router·worker, §6.2 보안, §6.3 transactional outbox)
  SECTION 14 Variance Bridge (§3 + "가정 변경" 레인; 결정→숫자 추적)
  SECTION 16 Reference Data 레이어 (외부지표→reference_data SCD2, 엔진 리팩터)
  SECTION 17 시간 기반 신뢰 감쇠 (§7.7: 노후 가정→stale→재확인)
  SECTION 18 콜드스타트/역사 백필 (§7.11)
  SECTION 20 거래처 Entity Resolution (§7.9: Fellegi-Sunter)
  SECTION 21 계약 개정 감지 (§7.8: 부속합의→새 버전·재스케줄)
  SECTION 22 민감도/What-if (§7.3: 드라이버 분기, 원장 불변)
  SECTION 23 지속 eval/회귀 (§6.9: 라우터·게이트 보정·그라운딩)
  SECTION 24 Variance 내러티브 생성 (§7.14: 그라운딩 — 환각 차단)
  SECTION 25 에이전트 관측/SLO (§6.7: 큐·게이트·충족·비용)
  SECTION 26 노트 분류·라우팅 (라이프사이클 §2: 단일 인박스→3갈래)
  SECTION 27 산출물 폴더링·매니페스트 (라이프사이클 §5: 택소노미·버전·발행)
  SECTION 28 외부 검토 반영 개선 (Splink u-추정·eval baseline/비용·예측 정확도)
  SECTION 29 구현 보강·정밀화 (ER EM/TF·하이브리드 검증·외부 어댑터·ABC·발송 sink·설정 기본값)
  SECTION 30 산출물 콘텐츠 스펙 (B4: 보고 덱/문서, grounded·house_style)
  SECTION 19 데모 / 실행 검증

실행:  python3 fixed_cost_decision_system.py
의존:  표준 라이브러리만 (numpy/pandas 불요). 회사 폐쇄망 드롭인.
"""
from __future__ import annotations
from dataclasses import dataclass, field, replace
from enum import IntEnum
from typing import Callable, Optional
import math, sqlite3, json, os


# ============================================================================
# SECTION 1 — 공통 토대
# ============================================================================
class Authority(IntEnum):                 # 출처권위 위계 (§7.6)
    SIGNED_CONTRACT = 5; INVOICE_PO = 4; BROKER_QUOTE = 3
    INTERNAL_ESTIMATE = 2; VERBAL_NOTE = 1


@dataclass
class Provenance:
    source_uri: str; authority: Authority; asof: str


@dataclass
class Input:
    name: str; value: float; unit: str; prov: Optional[Provenance] = None
    @property
    def sourced(self) -> bool: return self.prov is not None


@dataclass
class AnalysisResult:
    domain: str
    recommended_value: float                 # fcst로 흐르는 숫자
    recommendation: str                      # 행동 권고 → Task(승인)
    alternatives: list[str] = field(default_factory=list)
    confidence: float = 0.0
    conf_interval: tuple[float, float] = (0.0, 0.0)
    sensitivity: dict = field(default_factory=dict)
    materiality_band: str = "n/a"
    grounded: bool = False
    model_version: str = ""
    key_assumptions: dict = field(default_factory=dict)
    notes: list[str] = field(default_factory=list)

    def to_card(self, inputs: list[Input]) -> dict:
        return {
            "domain": self.domain,
            "inputs": [{"name": i.name, "value": i.value, "unit": i.unit,
                        "source_uri": i.prov.source_uri if i.prov else None,
                        "source_authority": int(i.prov.authority) if i.prov else None,
                        "asof": i.prov.asof if i.prov else None} for i in inputs],
            "model": self.domain, "model_version": self.model_version,
            "result": {"recommended_value": self.recommended_value,
                       "recommendation": self.recommendation, "alternatives": self.alternatives},
            "confidence": round(self.confidence, 3),
            "conf_interval": [round(self.conf_interval[0], 2), round(self.conf_interval[1], 2)],
            "sensitivity": self.sensitivity, "materiality_band": self.materiality_band,
            "grounded": int(self.grounded), "key_assumptions": self.key_assumptions}


def npv(rate: float, cfs: list[float]) -> float:
    return sum(cf / (1.0 + rate) ** t for t, cf in enumerate(cfs))

def pv_future(rate: float, cfs_y1: list[float]) -> float:    # t0 유출 없음 (VIU)
    return sum(cf / (1.0 + rate) ** t for t, cf in enumerate(cfs_y1, start=1))

def terminal_value_pv(last_cf, rate, growth, n):
    if rate <= growth: raise ValueError("r>g 필요")
    return (last_cf * (1.0 + growth) / (rate - growth)) / (1.0 + rate) ** n

def weighted_mean_se(values, weights):
    sw = sum(weights)
    if sw <= 0: raise ValueError("가중치 합 0")
    mean = sum(w * x for w, x in zip(weights, values)) / sw
    se = math.sqrt(sum((w ** 2) * (x - mean) ** 2 for w, x in zip(weights, values))) / sw
    n_eff = (sw ** 2) / sum(w ** 2 for w in weights)
    return mean, se, n_eff

def grounded_check(res: AnalysisResult, predicate: Callable[[AnalysisResult], bool]) -> bool:
    ok = bool(predicate(res)); res.grounded = ok
    if not ok: res.notes.append("GROUNDING FAIL: 권고가 수치와 불일치 → 검토 큐")
    return ok

def cap_confidence_by_provenance(base, inputs, material_names):
    notes, conf = [], base
    for nm in material_names:
        inp = next((i for i in inputs if i.name == nm), None)
        if inp is None or not inp.sourced:
            conf = min(conf, 0.5); notes.append(f"무출처 핵심입력 '{nm}' → 0.5 캡")
        elif inp.prov.authority <= Authority.INTERNAL_ESTIMATE:
            conf = min(conf, 0.7); notes.append(f"약한 출처 '{nm}' → 0.7 캡")
    return conf, notes

def materiality_band(amount, abs_threshold):
    a = abs(amount)
    if not abs_threshold: return "n/a"
    if a < abs_threshold * 0.5: return "below"
    if a < abs_threshold: return "near"
    return "material"

def confidence_from_stability(base_decision: bool, scenario_decisions, floor=0.5, ceiling=0.9):
    """결정이 민감도 시나리오에서 얼마나 안정적인가로 신뢰도 산출(임의 고정값 대체).
    근거: DCF/밸류에이션 신뢰도는 가정 섭동 하 결정 안정성에 비례(시나리오 일치율)."""
    if not scenario_decisions: return ceiling
    agree = sum(1 for d in scenario_decisions if bool(d) == base_decision) / len(scenario_decisions)
    return round(floor + (ceiling - floor) * agree, 3)


# ============================================================================
# SECTION 2 — 엔진 2.1 건물 임차 시장임대료 추정 + 유불리 (RICS 비교법)
# ============================================================================
RE_VERSION = "re_comparable/1.0.0"

@dataclass
class Subject:
    location: str; size_sqm: float; term_months: int; grade: int
    cost_basis: str; contract_rent_per_sqm: float

@dataclass
class Comp:
    comp_id: str; headline_rent_per_sqm: float; rent_free_months: float; term_months: int
    months_ago: float; size_sqm: float; grade: int; cost_basis: str
    same_location: bool; txn_type: str; prov: Provenance

@dataclass
class CompAdjustParams:
    market_trend_pm: float = 0.0; grade_step_pct: float = 0.06; size_elasticity: float = 0.03
    gross_to_net_opex_pct: float = 0.15; rf_halflife_months: float = 18.0
    letting_weight: float = 1.0; quoting_weight: float = 0.5

def _net_effective(c): 
    return c.headline_rent_per_sqm if c.term_months <= 0 else \
           c.headline_rent_per_sqm * (c.term_months - c.rent_free_months) / c.term_months

def _adjust(c, s, p):
    r = _net_effective(c)
    r *= (1.0 + p.market_trend_pm) ** c.months_ago
    r *= (1.0 + p.grade_step_pct) ** (s.grade - c.grade)
    if c.size_sqm > 0 and s.size_sqm > 0:
        r *= (1.0 - p.size_elasticity * math.log2(s.size_sqm / c.size_sqm))
    if c.cost_basis != s.cost_basis:
        if c.cost_basis == "gross" and s.cost_basis == "net": r *= (1.0 - p.gross_to_net_opex_pct)
        elif c.cost_basis == "net" and s.cost_basis == "gross": r *= (1.0 + p.gross_to_net_opex_pct)
    return r

def _weight(c, s, p):
    recency = math.exp(-c.months_ago / p.rf_halflife_months)
    grade_sim = 1.0 / (1.0 + abs(s.grade - c.grade))
    size_sim = 1.0 / (1.0 + abs(math.log2(max(c.size_sqm, 1) / max(s.size_sqm, 1))))
    loc = 1.0 if c.same_location else 0.6
    typ = p.letting_weight if c.txn_type == "letting" else p.quoting_weight
    return recency * grade_sim * size_sim * loc * typ

def estimate_market_rent(subject, comps, params=CompAdjustParams(),
                         abs_materiality=0.0, min_effective_comps=3.0, ci_level=0.80):
    if not comps: raise ValueError("비교 증거 없음")
    adj = [_adjust(c, subject, params) for c in comps]
    wts = [_weight(c, subject, params) for c in comps]
    mkt, se, n_eff = weighted_mean_se(adj, wts)
    z = {0.80: 1.2816, 0.90: 1.6449, 0.95: 1.9600}.get(round(ci_level, 2), 1.2816)
    ci = (mkt - z * se, mkt + z * se)   # 가중 비교 분산 기반 밴드(통계 보장 아님, ci_level 명시)
    fav_pct = (subject.contract_rent_per_sqm - mkt) / mkt
    annual_gap = fav_pct * mkt * subject.size_sqm * 12
    if fav_pct > 0.05:
        rec = f"재협상(목표 ₩{mkt:,.0f}/㎡, 연 절감 ~₩{abs(annual_gap):,.0f})"; alts = ["유지", "이전 검토"]
    elif fav_pct < -0.05:
        rec = "유지(시장 대비 유리)"; alts = ["재협상 불요"]
    else:
        rec = "시장 정합 — 현 조건 갱신"; alts = ["소폭 재협상"]
    base_conf = max(0.0, min(1.0, n_eff / (n_eff + 3.0)))
    inputs = [Input(f"comp:{c.comp_id}", c.headline_rent_per_sqm, "KRW/sqm/mo", c.prov) for c in comps]
    conf, cnotes = cap_confidence_by_provenance(base_conf, inputs, [f"comp:{comps[0].comp_id}"])
    res = AnalysisResult(domain="lease_favorability", recommended_value=mkt, recommendation=rec,
                         alternatives=alts, confidence=conf, conf_interval=ci,
                         sensitivity={"market_rent_band": [ci[0], ci[1]]},
                         materiality_band=materiality_band(annual_gap, abs_materiality),
                         model_version=RE_VERSION,
                         key_assumptions={"n_eff": round(n_eff, 2), "fav_pct": round(fav_pct, 4), "ci_level": ci_level},
                         notes=cnotes)
    if n_eff < min_effective_comps:
        res.notes.append(f"유효 비교 {n_eff:.1f}<{min_effective_comps} → 추가 증거 요청(신뢰도 제한)")
        res.confidence = min(res.confidence, 0.55)
    grounded_check(res, lambda r: ("재협상" in r.recommendation) == (fav_pct > 0.05) or abs(fav_pct) <= 0.05)
    return res, inputs


# ============================================================================
# SECTION 3 — 엔진 2.2 트럭 구매 vs 리스 (세후 DCF + IFRS16)
# ============================================================================
BVL_VERSION = "buy_vs_lease/1.0.0"

@dataclass
class OwnPlan:
    capex: float; life_years: int; residual_value: float
    annual_opex: float; tax_rate: float; disposal_cost: float = 0.0; in_service: int = 0

@dataclass
class LeasePlan:
    annual_payment: float; term_years: int; ibr: float
    escalation: float = 0.0; rent_free_months: float = 0.0; annual_service_in_payment: float = 0.0

def _own_cf(o: OwnPlan, horizon):
    dep = max(0.0, (o.capex - o.residual_value) / o.life_years)
    cf = [-o.capex]
    for t in range(1, horizon + 1):
        y = -o.annual_opex * (1 - o.tax_rate)
        if t <= o.life_years: y += dep * o.tax_rate
        cf.append(y)
    book_h = max(o.residual_value, o.capex - dep * min(horizon, o.life_years))
    gain = o.residual_value - book_h
    cf[horizon] += o.residual_value - gain * o.tax_rate - o.disposal_cost   # 처분손익 세금효과 대칭(손실=공제)
    return cf

def _lease_cf(l: LeasePlan, tax, horizon):
    cf = [0.0]
    for t in range(1, horizon + 1):
        pay = l.annual_payment * (1 + l.escalation) ** (t - 1) if t <= l.term_years else 0.0
        cf.append(-pay * (1 - tax))
    return cf

def lease_liability(l: LeasePlan):
    lease_only = l.annual_payment - l.annual_service_in_payment
    return sum(lease_only * (1 + l.escalation) ** (t - 1) / (1 + l.ibr) ** t
               for t in range(1, l.term_years + 1))

def _own_npv_res(o, rate, horizon, residual):
    return npv(rate, _own_cf(replace(o, residual_value=residual), horizon))

def _breakeven_residual(o, l, rate, horizon):
    ln = npv(rate, _lease_cf(l, o.tax_rate, horizon)); lo, hi = 0.0, o.capex
    for _ in range(60):
        mid = (lo + hi) / 2
        if _own_npv_res(o, rate, horizon, mid) < ln: lo = mid
        else: hi = mid
    return round((lo + hi) / 2, 0)

def net_advantage_to_lease(own: OwnPlan, lease: LeasePlan, kd_after_tax, horizon, salvage_rate=None):
    """리스-vs-구매(NAL): 리스는 부채등가 → 세후 리스료 + 포기 감가상각 세금절감을 세후 차입금리로
    할인. 잔존가치는 위험이 달라 별도(salvage_rate) 할인. NAL>0이면 리스 우위.
    근거: Myers·Dill·Bautista(1976), Ezzell·Miles(1983) 'after-tax cost of debt', Brealey·Myers·Allen.
    이자 세금절감은 별도 가산하지 않음(세후금리 할인이 이를 내포 — 이중계상 방지, Musumeci·O'Brien 2019)."""
    T = own.tax_rate; dep = max(0.0, (own.capex - own.residual_value) / own.life_years)
    sr = kd_after_tax if salvage_rate is None else salvage_rate
    pv_lease = 0.0
    for t in range(1, horizon + 1):
        pay = lease.annual_payment * (1 + lease.escalation) ** (t - 1) if t <= lease.term_years else 0.0
        dep_shield = dep * T if t <= own.life_years else 0.0
        pv_lease += (pay * (1 - T) + dep_shield) / (1 + kd_after_tax) ** t
    pv_salvage = own.residual_value / (1 + sr) ** horizon
    nal = own.capex - pv_lease - pv_salvage
    return {"nal": round(nal, 0), "kd_after_tax": round(kd_after_tax, 4),
            "decision": "리스" if nal > 0 else "구매", "pv_lease_outflow": round(pv_lease, 0)}

def analyze_buy_vs_lease(own, lease, hurdle_rate, horizon, abs_materiality=0.0,
                         prov_own=None, prov_lease=None):
    oc, lc = _own_cf(own, horizon), _lease_cf(lease, own.tax_rate, horizon)
    on, ln = npv(hurdle_rate, oc), npv(hurdle_rate, lc)
    diff = on - ln
    be = _breakeven_residual(own, lease, hurdle_rate, horizon)
    if diff > 0:
        rec = f"구매(소유 {on:,.0f}>리스 {ln:,.0f}, 우위 {diff:,.0f})"; alts = ["리스(현금유연성)"]
    else:
        rec = f"리스(리스 {ln:,.0f}>소유 {on:,.0f}, 우위 {-diff:,.0f})"; alts = ["구매(장기·고주행)"]
    nal = net_advantage_to_lease(own, lease, kd_after_tax=lease.ibr * (1 - own.tax_rate), horizon=horizon)
    sens = {"hurdle_+1pp": [npv(hurdle_rate + 0.01, oc), npv(hurdle_rate + 0.01, lc)],
            "residual_-20pct": [_own_npv_res(own, hurdle_rate, horizon, own.residual_value * 0.8), ln],
            "breakeven_residual": be, "NAL_equivalent_loan": nal["nal"]}
    inputs = [Input("capex", own.capex, "KRW", prov_own), Input("annual_lease", lease.annual_payment, "KRW", prov_lease)]
    # 신뢰도: 구매/리스 결정이 섭동(허들±·잔존-20%·NAL) 하 안정적인가
    base_buy = diff > 0
    scen = [sens["hurdle_+1pp"][0] - sens["hurdle_+1pp"][1] > 0,
            sens["residual_-20pct"][0] - sens["residual_-20pct"][1] > 0,
            nal["nal"] <= 0]   # NAL>0=리스 우위 → 구매결정과 반대
    conf_bvl = confidence_from_stability(base_buy, scen)
    res = AnalysisResult(domain="buy_vs_lease", recommended_value=max(on, ln), recommendation=rec,
                         alternatives=alts, confidence=conf_bvl, conf_interval=(min(on, ln), max(on, ln)),
                         sensitivity=sens, materiality_band=materiality_band(diff, abs_materiality),
                         model_version=BVL_VERSION,
                         key_assumptions={"hurdle": hurdle_rate, "ibr": lease.ibr, "horizon": horizon},
                         notes=[f"IFRS16 리스부채(표시)=₩{lease_liability(lease):,.0f}",
                                f"의사결정 할인=허들 {hurdle_rate:.1%}(자본예산 스크린, ≠IBR {lease.ibr:.1%})",
                                f"NAL(등가대출, 세후차입 {lease.ibr*(1-own.tax_rate):.1%}): ₩{nal['nal']:,.0f} → {nal['decision']} "
                                f"— 재무이론상 리스=부채등가 결정; 할인율 선택이 결론을 바꿀 수 있음(두 관점 병기)"])
    grounded_check(res, lambda r: ("구매" in r.recommendation) == (diff > 0))
    return res, inputs


# ============================================================================
# SECTION 4 — 엔진 2.3 IAS 36 손상 (VIU + CGU)
# ============================================================================
IMP_VERSION = "ias36_impairment/1.0.0"

@dataclass
class CGUAsset:
    asset_id: str; carrying: float; is_goodwill: bool = False
    life_years: int = 0; elapsed_years: int = 0; residual: float = 0.0   # 투영 배선용
    recoverable_floor: float = 0.0   # IAS36 ¶105: max(FVLCD,VIU,0) 미만 감액 금지

@dataclass
class CGU:
    cgu_id: str; assets: list[CGUAsset]
    pretax_cashflows_y1_n: list[float]; pretax_rate: float
    terminal_growth: float = 0.0; use_terminal: bool = True
    fair_value: float = 0.0; costs_of_disposal: float = 0.0
    @property
    def carrying_total(self): return sum(a.carrying for a in self.assets)

def value_in_use(c: CGU):
    n = len(c.pretax_cashflows_y1_n)
    viu = pv_future(c.pretax_rate, c.pretax_cashflows_y1_n)
    # IAS36 ¶33: 명시기간은 유용수명 내, 성장률은 장기평균 초과 불가 → growth<rate일 때만 잔존가치 가산
    if c.use_terminal and n > 0 and c.terminal_growth < c.pretax_rate:
        viu += terminal_value_pv(c.pretax_cashflows_y1_n[-1], c.pretax_rate, c.terminal_growth, n)
    return viu

def fvlcd(c: CGU): return max(0.0, c.fair_value - c.costs_of_disposal)

def allocate_loss(c: CGU, loss):
    """IAS36 ¶104-105: 영업권 먼저 → 나머지 장부금액 비례. 단, 개별 자산은 max(FVLCD,VIU,0)
    미만으로 감액 불가(¶105 바닥값); 바닥에 닿아 배분 못한 초과분은 나머지 자산에 재배분(반복)."""
    alloc = {a.asset_id: 0.0 for a in c.assets}
    rem = loss
    for a in [x for x in c.assets if x.is_goodwill]:          # 1) 영업권
        take = min(a.carrying, rem); alloc[a.asset_id] = take; rem -= take
    others = [x for x in c.assets if not x.is_goodwill]
    cap = {a.asset_id: max(0.0, a.carrying - max(a.recoverable_floor, 0.0)) for a in others}  # 흡수 한도
    gw_taken = sum(alloc[g.asset_id] for g in c.assets if g.is_goodwill)
    for _ in range(len(others) + 2):                          # 2) 비례 배분 + 초과분 재배분(반복)
        active = [a for a in others if (cap[a.asset_id] - alloc[a.asset_id]) > 1e-6]
        rem = (loss - gw_taken) - sum(alloc[a.asset_id] for a in others)
        if rem <= 1e-6 or not active: break
        base = sum(a.carrying for a in active)
        for a in active:
            take = min(rem * (a.carrying / base), cap[a.asset_id] - alloc[a.asset_id])
            alloc[a.asset_id] += take
    return [(aid, round(v, 0)) for aid, v in alloc.items() if v > 0]

def test_impairment(c: CGU, abs_materiality=0.0, prov_fv=None):
    viu, fv = value_in_use(c), fvlcd(c)
    recoverable = max(viu, fv); carrying = c.carrying_total
    loss = max(0.0, carrying - recoverable); headroom = recoverable - carrying
    basis = "VIU" if viu >= fv else "FVLCD"
    if loss > 0:
        alloc = allocate_loss(c, loss)
        rec = f"손상 인식 ₩{loss:,.0f} (회수가능 {recoverable:,.0f}<{carrying:,.0f}, 기준={basis})"; alts = ["CGU 재검토"]
    else:
        alloc = []; rec = f"손상 없음 (headroom ₩{headroom:,.0f}, 기준={basis})"; alts = ["환입 검토(영업권 제외)"]
    def vw(rate=None, g=None, s=1.0):
        cc = replace(c, pretax_rate=c.pretax_rate if rate is None else rate,
                     terminal_growth=c.terminal_growth if g is None else g,
                     pretax_cashflows_y1_n=[x * s for x in c.pretax_cashflows_y1_n])
        return max(value_in_use(cc), fv) - carrying
    sens = {"pretax_+1pp": round(vw(rate=c.pretax_rate + 0.01), 0),
            "growth_-1pp": round(vw(g=c.terminal_growth - 0.01), 0),
            "cashflow_-10pct": round(vw(s=0.9), 0)}
    # 신뢰도: 손상 여부 결정이 민감도 섭동 하 안정적인가(headroom<0 ⇒ 손상)
    base_impair = loss > 0
    conf = confidence_from_stability(base_impair, [v < 0 for v in sens.values()])
    conf, cnotes = cap_confidence_by_provenance(conf, [Input("fair_value", c.fair_value, "KRW", prov_fv)], ["fair_value"])
    ka = {"pretax_rate": c.pretax_rate, "terminal_growth": c.terminal_growth,
          "explicit_years": len(c.pretax_cashflows_y1_n), "basis": basis,
          "fvlcd": round(fv, 0), "viu": round(viu, 0), "loss": round(loss, 0)}
    inputs = [Input("fair_value", c.fair_value, "KRW", prov_fv), Input("carrying", carrying, "KRW", None)]
    res = AnalysisResult(domain="impairment", recommended_value=loss, recommendation=rec,
                         alternatives=alts, confidence=conf,
                         conf_interval=(min(0.0, headroom), max(0.0, headroom)),
                         sensitivity=sens, materiality_band=materiality_band(loss if loss else headroom, abs_materiality),
                         model_version=IMP_VERSION, key_assumptions=ka,
                         notes=[f"loss_allocation={alloc}", "CGU 경계 사람 승인 필수(상향집계 편향·SOX)", *cnotes])
    grounded_check(res, lambda r: ("손상 인식" in r.recommendation) == (carrying > recoverable))
    res.notes.append(f"alloc_detail={alloc}")
    res._alloc = alloc      # 투영 배선에서 사용
    return res, inputs


# ============================================================================
# SECTION 5 — 투영 엔진 §7.1 (상각 roll-forward · 리스 step-up · run-rate)
# ============================================================================
PROJ_VERSION = "projection/1.0.0"

def depreciation_schedule(cost, residual, life_years, start_period=1):
    """직선 상각 연간 스케줄. (period, depreciation, closing_carrying)."""
    annual = max(0.0, (cost - residual) / life_years)
    rows, carry = [], cost
    for i in range(life_years):
        carry = max(residual, carry - annual)
        rows.append((start_period + i, round(annual, 0), round(carry, 0)))
    return rows

def revised_depreciation_after_impairment(asset: CGUAsset, loss_alloc: float, start_period=1):
    """IAS 36: 손상 후 수정 장부금액을 잔여 내용연수로 상각."""
    revised_carrying = asset.carrying - loss_alloc
    remaining = max(1, asset.life_years - asset.elapsed_years)
    return depreciation_schedule(revised_carrying, asset.residual, remaining, start_period), revised_carrying

def lease_schedule(l: LeasePlan, start_period=1):
    """IFRS16: 리스부채 상환(이자/원금) + ROU 직선상각 + P&L(이자+ROU상각)."""
    liability = lease_liability(l); rou0 = liability; rou_dep = rou0 / l.term_years
    rows, liab = [], liability
    for t in range(1, l.term_years + 1):
        pay = (l.annual_payment - l.annual_service_in_payment) * (1 + l.escalation) ** (t - 1)
        interest = liab * l.ibr; principal = pay - interest; liab = liab - principal
        pnl = interest + rou_dep
        rows.append((start_period + t - 1, round(pay, 0), round(interest, 0),
                     round(principal, 0), round(max(0.0, liab), 0), round(rou_dep, 0), round(pnl, 0)))
    return rows  # (period, payment, interest, principal, liability_end, rou_dep, pnl_charge)

def runrate_schedule(remaining_commitment, periods, start_period=1):
    per = remaining_commitment / periods; rem = remaining_commitment; rows = []
    for i in range(periods):
        rem = max(0.0, rem - per)
        rows.append((start_period + i, round(per, 0), round(rem, 0)))
    return rows  # (period, run_rate, remaining_commitment)


# ============================================================================
# SECTION 6 — 의사결정 → 투영 배선 (결정이 fcst 라인 스케줄로)
# ============================================================================
def project_fcst_lines(domain, *, own: OwnPlan = None, lease: LeasePlan = None,
                       recommendation: str = "", cgu: CGU = None, alloc: list = None,
                       renewal_payment: float = None, renewal_term: int = None):
    """결정 도메인·권고에 따라 fcst 라인 투영. (line_type, schedule, note) 리스트 반환."""
    out = []
    if domain == "buy_vs_lease":
        if "구매" in recommendation and own:
            out.append(("depreciation", depreciation_schedule(own.capex, own.residual_value, own.life_years), "소유: 직선상각"))
        elif lease:
            out.append(("lease_pnl", lease_schedule(lease), "리스: ROU상각+이자(IFRS16)"))
    elif domain == "impairment" and cgu and alloc:
        amap = {aid: amt for aid, amt in alloc}
        for a in cgu.assets:
            if a.is_goodwill or a.life_years <= 0: continue
            sched, revised = revised_depreciation_after_impairment(a, amap.get(a.asset_id, 0.0))
            out.append((f"dep_revised:{a.asset_id}", sched, f"손상후 수정장부 ₩{revised:,.0f}/잔여 {a.life_years-a.elapsed_years}년"))
    elif domain == "lease_favorability" and renewal_payment and renewal_term:
        # 갱신 가정(reasonably certain) → 갱신 리스 스케줄로 투영
        rl = LeasePlan(annual_payment=renewal_payment, term_years=renewal_term, ibr=0.055)
        out.append(("lease_renewal_pnl", lease_schedule(rl), "갱신 가정 반영 리스 스케줄"))
    return out


# ============================================================================
# SECTION 7 — SOX/ICFR 통제 매트릭스 §7.13 (증적 자동수집)
# ============================================================================
# COSO/PCAOB AS 2201 관점: 고정비 핵심통제와 운영 유효성 증거.
CONTROL_MATRIX = [
    # control_id, risk, frequency, owner
    ("C-IMP-01", "손상 측정 오류(IAS 36)", "quarterly", "fixedcost-fpna"),
    ("C-IMP-02", "CGU 상향집계 편향(손상 은폐)", "annual/on-change", "controller"),
    ("C-LSE-01", "리스 회계 완전성·IBR(IFRS 16)", "on-new-lease", "treasury+fpna"),
    ("C-DEP-01", "감가상각 정책 적정성(IAS 16)", "annual", "controller"),
    ("C-SOD-01", "직무분리(요청자≠승인자)", "per-transaction", "fpna-lead"),
    ("C-CHG-01", "변경관리(방법/플레이북 SemVer+eval)", "per-change", "fpna-lead"),
]
DOMAIN_CONTROLS = {
    "impairment": ["C-IMP-01", "C-IMP-02", "C-SOD-01"],
    "buy_vs_lease": ["C-LSE-01", "C-DEP-01", "C-SOD-01"],
    "lease_favorability": ["C-LSE-01", "C-SOD-01"],
}

def init_controls(con):
    con.executescript("""
      CREATE TABLE IF NOT EXISTS control_matrix(control_id TEXT PRIMARY KEY, risk TEXT, frequency TEXT, owner TEXT);
      CREATE TABLE IF NOT EXISTS control_evidence(
        id INTEGER PRIMARY KEY, control_id TEXT, analysis_id INTEGER, domain TEXT,
        key_assumptions TEXT, approval_status TEXT, grounded INTEGER, at TEXT);
    """)
    con.executemany("INSERT OR IGNORE INTO control_matrix VALUES(?,?,?,?)", CONTROL_MATRIX)
    con.commit()

def collect_control_evidence(con, analysis_id, res: AnalysisResult, approval_status, now):
    for cid in DOMAIN_CONTROLS.get(res.domain, []):
        con.execute("""INSERT INTO control_evidence(control_id,analysis_id,domain,key_assumptions,approval_status,grounded,at)
                       VALUES(?,?,?,?,?,?,?)""",
                    (cid, analysis_id, res.domain, json.dumps(res.key_assumptions, ensure_ascii=False),
                     approval_status, int(res.grounded), now))
    con.commit()

def icfr_summary(con):
    """통제별 최신 증거·운영 유효성. effective = 증거존재 ∧ 전부 grounded ∧ 전부 승인완료."""
    rows = []
    for cid, risk, freq, owner in con.execute("SELECT control_id,risk,frequency,owner FROM control_matrix"):
        ev = con.execute(
            """SELECT COUNT(*), MAX(at), SUM(grounded),
                      SUM(CASE WHEN approval_status='approved' THEN 1 ELSE 0 END)
               FROM control_evidence WHERE control_id=?""", (cid,)).fetchone()
        cnt, last, grounded_sum, approved_sum = ev[0], ev[1], (ev[2] or 0), (ev[3] or 0)
        if not cnt:
            status = "no-evidence"
        elif grounded_sum == cnt and approved_sum == cnt:
            status = "effective"
        elif approved_sum < cnt:
            status = "pending-approval"
        else:
            status = "review"
        rows.append((cid, risk, freq, cnt, last or "-", status))
    return rows


# ============================================================================
# SECTION 8 — Decision Analysis 카드 + 게이트 + SQLite
# ============================================================================
DB_SCHEMA = """
CREATE TABLE IF NOT EXISTS decision_analysis(
  id INTEGER PRIMARY KEY, question TEXT, domain TEXT, fcst_line TEXT, evidence_kind TEXT, inputs TEXT, model TEXT, model_version TEXT,
  result TEXT, confidence REAL, conf_interval TEXT, sensitivity TEXT, materiality_band TEXT,
  grounded INTEGER, status TEXT, linked_assumption_card TEXT, next_review TEXT, created_at TEXT);
CREATE TABLE IF NOT EXISTS task_card(
  id INTEGER PRIMARY KEY, analysis_id INTEGER, task_type TEXT, payload TEXT, risk_tier TEXT,
  status TEXT, requester TEXT, approver TEXT, executed_result TEXT, created_at TEXT);
CREATE TABLE IF NOT EXISTS fcst_line_projection(
  id INTEGER PRIMARY KEY, analysis_id INTEGER, line_type TEXT, period INTEGER, value REAL, note TEXT);
CREATE TABLE IF NOT EXISTS decision_register(decision_id TEXT, analysis_id INTEGER, event TEXT, actor TEXT, at TEXT);
CREATE TABLE IF NOT EXISTS work_item(
  id INTEGER PRIMARY KEY, dedup_key TEXT UNIQUE, source TEXT, raw_ref TEXT,
  correlation_token TEXT, payload TEXT, trust TEXT, status TEXT, attempts INTEGER DEFAULT 0, created_at TEXT);
CREATE TABLE IF NOT EXISTS request_register(
  request_id TEXT PRIMARY KEY, fcst_line TEXT, owner TEXT, due_date TEXT,
  correlation_token TEXT, status TEXT, created_at TEXT);
CREATE TABLE IF NOT EXISTS outbox(
  id INTEGER PRIMARY KEY, task_id INTEGER, side_effect TEXT, idempotency_key TEXT UNIQUE,
  status TEXT, external_ref TEXT, attempts INTEGER DEFAULT 0, created_at TEXT);
CREATE TABLE IF NOT EXISTS variance_lane(
  id INTEGER PRIMARY KEY, line_id TEXT, scenario_pair TEXT, lane TEXT, amount REAL, evidence TEXT, created_at TEXT);
CREATE TABLE IF NOT EXISTS ref_snapshot(
  ref_set TEXT, snapshot_date TEXT, source TEXT, source_url TEXT, license TEXT,
  version TEXT, sha256 TEXT, ingested_at TEXT, PRIMARY KEY(ref_set, version));
CREATE TABLE IF NOT EXISTS regional_rent_benchmark(
  region TEXT, property_type TEXT, grade INTEGER, period TEXT,
  rent_per_sqm REAL, rent_index REAL, vacancy REAL,
  source_version TEXT, valid_from TEXT, valid_to TEXT);
CREATE TABLE IF NOT EXISTS ibr_matrix(
  currency TEXT, term_band TEXT, security TEXT, ibr REAL, benchmark TEXT, spread REAL,
  source_version TEXT, valid_from TEXT, valid_to TEXT);
CREATE TABLE IF NOT EXISTS vendor_cluster(
  raw_name TEXT, biz_no TEXT, canonical_vendor_id TEXT, match_score REAL, method TEXT, status TEXT, created_at TEXT);
CREATE TABLE IF NOT EXISTS contract_master(
  contract_id TEXT PRIMARY KEY, counterparty TEXT, asset_or_property TEXT, contract_no TEXT,
  version INTEGER, amendment_seq INTEGER, supersedes_contract_id TEXT,
  monthly_amount REAL, term_months INTEGER, escalation REAL, start_date TEXT, end_date TEXT,
  status TEXT, source_doc TEXT, fcst_line TEXT, created_at TEXT);
CREATE TABLE IF NOT EXISTS sensitivity_case(
  case_id TEXT, base_ref TEXT, domain TEXT, overrides TEXT, metric TEXT, value REAL, delta_vs_base REAL, flips INTEGER, created_at TEXT);
CREATE TABLE IF NOT EXISTS rebuild_request(
  id INTEGER PRIMARY KEY, trigger TEXT, affected TEXT, status TEXT, created_at TEXT);
CREATE TABLE IF NOT EXISTS playbook_card(
  playbook_id TEXT PRIMARY KEY, trigger TEXT, status TEXT, proposed_for TEXT, created_at TEXT);
CREATE TABLE IF NOT EXISTS eval_run(
  id INTEGER PRIMARY KEY, suite TEXT, metric TEXT, value REAL, threshold REAL, passed INTEGER, detail TEXT, run_at TEXT);
CREATE TABLE IF NOT EXISTS ops_run(
  run_id TEXT PRIMARY KEY, job TEXT, started_at TEXT, finished_at TEXT, status TEXT, tokens INTEGER, cost REAL, facets TEXT);
CREATE TABLE IF NOT EXISTS note_register(
  note_id TEXT PRIMARY KEY, type TEXT, topic TEXT, fcst_line TEXT, sensitivity TEXT,
  route TEXT, confidence REAL, tags TEXT, status TEXT, captured_at TEXT);
CREATE TABLE IF NOT EXISTS artifact_register(
  artifact_id TEXT PRIMARY KEY, period TEXT, audience TEXT, atype TEXT, scope TEXT,
  version INTEGER, run_id TEXT, path TEXT, ext TEXT, status TEXT, manifest TEXT, created_at TEXT);
CREATE TABLE IF NOT EXISTS forecast_actual(
  line TEXT, period TEXT, forecast REAL, actual REAL, recorded_at TEXT);
CREATE TABLE IF NOT EXISTS sent_log(
  id INTEGER PRIMARY KEY, task_id INTEGER, idempotency_key TEXT UNIQUE, channel TEXT, detail TEXT, sent_at TEXT);
"""
DOMAIN_TASK = {
    "lease_favorability": ("propose_renegotiation", "med"),
    "buy_vs_lease": ("propose_capex_or_lease", "med"),
    "impairment": ("confirm_impairment", "high"),
}

# ── 회계 이벤트 + 회계팀 통지 정책 ─────────────────────────────────────────────
# 의사결정/계약이벤트는 통상 회계처리(회계팀 통지사항)와 연결된다. 기본 동작은 *포캐스트 반영*이고,
# 회계팀 통지는 이벤트 유형별로 기본 OFF. 필요할 때 하나씩 켠다(런타임 set 또는 env FPNA_NOTIFY_<EVENT>).
STD_REF = {"lease_recognition": "IFRS 16(리스 최초인식: ROU·리스부채)",
           "lease_remeasurement": "IFRS 16 ¶40-46(리스 변경·재측정)",
           "asset_acquisition": "IAS 16(유형자산 취득·감가)",
           "impairment_loss": "IAS 36(손상차손·수정 감가)"}
ACCOUNTING_NOTIFY = {}   # 명시적 in-process 오버라이드(런타임/테스트). 없으면 env, 그래도 없으면 False(OFF)

def is_accounting_notify_on(event):
    if event in ACCOUNTING_NOTIFY: return bool(ACCOUNTING_NOTIFY[event])
    return os.environ.get(f"FPNA_NOTIFY_{event.upper()}", "0") == "1"   # 운영 영속 토글(재시작 유지)

def set_accounting_notify(event, on=True):
    """회계팀 통지를 이벤트 유형별로 켜고 끈다(기본 OFF). 예: set_accounting_notify('impairment_loss', True)."""
    ACCOUNTING_NOTIFY[event] = bool(on); return {"event": event, "on": bool(on)}

def accounting_event_for(domain, recommendation=""):
    """의사결정 도메인(+결과) → 회계 이벤트 유형. 없으면 None(회계 이벤트 아님)."""
    if domain == "impairment": return "impairment_loss"
    if domain == "lease_favorability": return "lease_remeasurement"
    if domain == "buy_vs_lease":
        return "lease_recognition" if "리스" in (recommendation or "") else "asset_acquisition"
    return None

def maybe_notify_accounting(con, analysis_id, domain, recommendation="", fcst_line=None, amount=None, now="2026-06-14"):
    """회계 이벤트면 통지 정책 확인 → ON일 때만 notify_accounting Task 생성(사람 승인 후 발송).
    OFF(기본)면 아무것도 안 함 — 포캐스트 반영만. 반환: 이벤트 유형 + 통지 여부."""
    ev = accounting_event_for(domain, recommendation)
    if not ev or not is_accounting_notify_on(ev):
        return {"accounting_event": ev, "notified": False}
    con.execute("""INSERT INTO task_card(analysis_id,task_type,payload,risk_tier,status,requester,approver,created_at)
        VALUES(?,'notify_accounting',?, 'low','pending_approval','system',NULL,?)""",
        (analysis_id, json.dumps({"event": ev, "fcst_line": fcst_line, "summary": recommendation,
                                  "amount": amount, "standard": STD_REF.get(ev, "")}, ensure_ascii=False), now))
    con.commit()
    return {"accounting_event": ev, "notified": True}

def init_db(path=":memory:"):
    con = sqlite3.connect(path); con.executescript(DB_SCHEMA); init_controls(con); return con

EVIDENCE_KIND = {"lease_favorability": "market_estimate", "buy_vs_lease": "market_estimate", "impairment": "projection"}

def gate_decision(grounded, confidence, materiality_band, tier, conf_threshold=0.6):
    """자동반영 게이트(§6.1/§6.8) — eval 보정에서 재사용하기 위해 순수 함수로 분리."""
    return bool(grounded and confidence >= conf_threshold
                and materiality_band != "material" and tier != "high")

def submit(con, question, res: AnalysisResult, inputs, *, projections=None,
           fcst_line=None, evidence_kind=None, conf_threshold=0.6, now="2026-06-14"):
    card = res.to_card(inputs)
    ek = evidence_kind or EVIDENCE_KIND.get(res.domain, "market_estimate")
    task_type, tier = DOMAIN_TASK.get(res.domain, ("propose_action", "med"))
    auto_ok = gate_decision(res.grounded, res.confidence, res.materiality_band, tier, conf_threshold)
    status = "applied" if auto_ok else "review"
    if tier == "high": res.notes.append("high-tier → 분석 숫자 자동반영 불가(검토 강제, SOX)")
    cur = con.execute("""INSERT INTO decision_analysis(question,domain,fcst_line,evidence_kind,inputs,model,model_version,result,
            confidence,conf_interval,sensitivity,materiality_band,grounded,status,linked_assumption_card,next_review,created_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (question, card["domain"], fcst_line, ek, json.dumps(card["inputs"], ensure_ascii=False), card["model"],
         card["model_version"], json.dumps(card["result"], ensure_ascii=False), card["confidence"],
         json.dumps(card["conf_interval"]), json.dumps(card["sensitivity"], ensure_ascii=False),
         card["materiality_band"], card["grounded"], status,
         f"assumption::{card['domain']}" if auto_ok else None, now, now))
    aid = cur.lastrowid
    con.execute("""INSERT INTO task_card(analysis_id,task_type,payload,risk_tier,status,requester,approver,created_at)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (aid, task_type, json.dumps({"recommendation": res.recommendation, "value": res.recommended_value}, ensure_ascii=False),
                 tier, "pending_approval", "fpna-analyst", None, now))
    # fcst 라인 투영 적재 (§7.1 배선)
    n_proj = 0
    for line_type, sched, note in (projections or []):
        for row in sched:
            period = row[0]
            value = row[6] if line_type.endswith("pnl") else row[1]   # 리스=P&L charge, 상각=dep
            con.execute("INSERT INTO fcst_line_projection(analysis_id,line_type,period,value,note) VALUES(?,?,?,?,?)",
                        (aid, line_type, period, value, note)); n_proj += 1
    # SOX 통제 증적 자동수집 (§7.13)
    collect_control_evidence(con, aid, res, approval_status="pending_approval", now=now)
    con.execute("INSERT INTO decision_register VALUES(?,?,?,?,?)", (f"D-{aid}", aid, "analyzed", "worker", now))
    con.commit()
    # 회계 이벤트면 통지(기본 OFF; 토글 ON일 때만 notify_accounting Task) — 기본은 위 포캐스트 반영뿐
    notify = maybe_notify_accounting(con, aid, res.domain, res.recommendation, fcst_line, res.recommended_value, now)
    return {"analysis_id": aid, "status": status, "auto_applied": auto_ok, "grounded": res.grounded,
            "confidence": round(res.confidence, 3), "materiality": res.materiality_band,
            "task": f"{task_type}({tier})→pending_approval", "fcst_rows": n_proj, "recommendation": res.recommendation,
            "accounting_event": notify["accounting_event"], "notify_accounting": notify["notified"]}


# ============================================================================
# SECTION 10 — 승인 워크플로 (§6.8: 직무분리 SoD + 실행 + ICFR 완성)
# ============================================================================
class SoDViolation(Exception): pass

def approve_task(con, task_id, approver, now="2026-06-14"):
    """Task 승인 → outbox 적재. SoD(요청자≠승인자) 강제. 실제 실행은 process_outbox(멱등)."""
    row = con.execute("SELECT analysis_id,task_type,risk_tier,status,requester FROM task_card WHERE id=?",
                      (task_id,)).fetchone()
    if not row: raise ValueError(f"task {task_id} 없음")
    analysis_id, task_type, tier, status, requester = row
    if status != "pending_approval": raise ValueError(f"승인 불가 상태: {status}")
    if approver == requester:                              # 직무분리(C-SOD-01)
        raise SoDViolation(f"SoD 위반: 요청자={requester} == 승인자={approver}")
    con.execute("UPDATE task_card SET status='approved', approver=? WHERE id=?", (approver, task_id))
    # Transactional outbox(§6.3): 부작용 의도를 멱등키로 기록. 실행은 분리.
    enqueue_outbox(con, task_id, side_effect=task_type, idem_key=f"task:{task_id}", now=now)
    con.execute("INSERT INTO decision_register VALUES(?,?,?,?,?)",
                (f"D-{analysis_id}", analysis_id, "approved", approver, now))
    con.commit()
    return {"task_id": task_id, "status": "approved", "approver": approver, "sod_ok": True, "outboxed": task_type}

def reject_task(con, task_id, approver, reason="", now="2026-06-14"):
    """Task 거부(SoD: 거부자≠요청자). pending_approval→rejected(종결). 연계 결정은 review 유지(부작용 없음)."""
    row = con.execute("SELECT status,requester FROM task_card WHERE id=?", (task_id,)).fetchone()
    if not row: raise ValueError(f"task {task_id} 없음")
    status, requester = row
    if status != "pending_approval": raise ValueError(f"거부 불가 상태: {status}")
    if approver == requester: raise SoDViolation(f"SoD 위반: 요청자={requester} == 거부자={approver}")
    con.execute("UPDATE task_card SET status='rejected', approver=?, executed_result=? WHERE id=?",
                (approver, json.dumps({"rejected": True, "reason": reason}, ensure_ascii=False), task_id))
    con.execute("INSERT INTO decision_register VALUES(?,?,?,?,?)",
                (f"D-{row[0] if False else task_id}", None, "rejected", approver, now))
    con.commit()
    return {"task_id": task_id, "status": "rejected", "reason": reason, "by": approver}


# ============================================================================
# SECTION 11 — openpyxl 보고서 (View Contract 적용; xlsx 스킬 준수)
# ============================================================================
# 적용 규약: R1 전체 시간축 ruler · R3 recon 블록(SUM 수식) · evidence health 컬럼.
# 서식: Arial, 통화 $#,##0(₩), 음수 괄호, 합계는 하드코딩 아닌 수식.
def build_report(con, path, periods=8):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR = Font(name="Arial", bold=True, color="FFFFFF")
    BOLD = Font(name="Arial", bold=True)
    BODY = Font(name="Arial")
    BLUE = Font(name="Arial", color="0000FF")          # 입력(하드코딩 데이터)
    NAVY = PatternFill("solid", fgColor="1F3864")
    GREY = PatternFill("solid", fgColor="D9D9D9")
    YEL = PatternFill("solid", fgColor="FFFF00")
    thin = Side(style="thin", color="BFBFBF"); BORD = Border(*([thin] * 4))
    KRW = u'₩#,##0;(₩#,##0);"-"'

    wb = Workbook()

    # ---- Sheet 1: Fixed Cost Forecast (전체 시간축 ruler + recon) ----
    ws = wb.active; ws.title = "FixedCost_Forecast"
    ws["A1"] = "고정비 Forecast (계약 중심 결정 반영)"; ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws["A2"] = f"기간 ruler 1..{periods} (데이터 없는 기간도 0 유지 — View Contract R1)"; ws["A2"].font = BODY

    hdr_row = 4
    headers = ["analysis_id", "line_type", "근거(note)", "evidence_health"] + [f"P{p}" for p in range(1, periods + 1)] + ["row_total"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(hdr_row, c, h); cell.font = HDR; cell.fill = NAVY; cell.alignment = Alignment(horizontal="center"); cell.border = BORD
    # 라인 = (analysis_id,line_type) 그룹. 전체 ruler로 피벗(없는 기간=0).
    keys = list(con.execute("SELECT DISTINCT analysis_id,line_type,note FROM fcst_line_projection ORDER BY analysis_id,line_type"))
    health = {r[0]: r[1] for r in con.execute("SELECT id, (CASE WHEN grounded=1 THEN 'grounded' ELSE 'UNGROUNDED' END)||'/'||status FROM decision_analysis")}
    first_pcol = 5
    r = hdr_row + 1
    for aid, lt, note in keys:
        vals = {p: 0.0 for p in range(1, periods + 1)}
        for period, value in con.execute("SELECT period,value FROM fcst_line_projection WHERE analysis_id=? AND line_type=?", (aid, lt)):
            if 1 <= period <= periods: vals[period] = value
        ws.cell(r, 1, aid).font = BODY
        ws.cell(r, 2, lt).font = BODY
        ws.cell(r, 3, note).font = BODY
        hc = ws.cell(r, 4, health.get(aid, "n/a")); hc.font = BODY
        if "UNGROUNDED" in hc.value or "review" in hc.value: hc.fill = YEL   # 주의 강조
        for p in range(1, periods + 1):
            cell = ws.cell(r, first_pcol + p - 1, vals[p]); cell.font = BLUE; cell.number_format = KRW; cell.border = BORD
        # row_total = 수식(하드코딩 금지)
        a1 = f"{get_column_letter(first_pcol)}{r}"; a2 = f"{get_column_letter(first_pcol + periods - 1)}{r}"
        tot = ws.cell(r, first_pcol + periods, f"=SUM({a1}:{a2})"); tot.font = BOLD; tot.number_format = KRW; tot.border = BORD
        r += 1

    # ---- recon 블록 (R3): 기간 합계 + tie-out(열합 vs 행합) 수식 ----
    rr = r + 1
    ws.cell(rr, 3, "RECON: 기간 합계").font = BOLD; ws.cell(rr, 3).fill = GREY
    for p in range(1, periods + 1):
        col = get_column_letter(first_pcol + p - 1)
        cell = ws.cell(rr, first_pcol + p - 1, f"=SUM({col}{hdr_row+1}:{col}{r-1})")
        cell.font = BOLD; cell.number_format = KRW; cell.fill = GREY
    gt_col = get_column_letter(first_pcol + periods)
    grand = ws.cell(rr, first_pcol + periods, f"=SUM({gt_col}{hdr_row+1}:{gt_col}{r-1})"); grand.font = BOLD; grand.number_format = KRW; grand.fill = GREY
    # tie-out: 기간합계들의 합 == 행합계들의 합 (0이어야 정상)
    tr = rr + 1
    ws.cell(tr, 3, "RECON: tie-out (열합−행합, =0)").font = BOLD
    pstart = get_column_letter(first_pcol); pend = get_column_letter(first_pcol + periods - 1)
    tie = ws.cell(tr, first_pcol, f"=SUM({pstart}{rr}:{pend}{rr})-{gt_col}{rr}"); tie.font = BOLD; tie.number_format = KRW
    for col_i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 16
    ws.column_dimensions["C"].width = 34

    # ---- Sheet 2: Decision Register ----
    ws2 = wb.create_sheet("Decision_Register")
    h2 = ["id", "domain", "recommendation", "confidence", "materiality", "grounded", "status", "model_version"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(1, c, h); cell.font = HDR; cell.fill = NAVY
    rr2 = 2
    for row in con.execute("""SELECT id,domain,json_extract(result,'$.recommendation'),confidence,materiality_band,grounded,status,model_version
                              FROM decision_analysis ORDER BY id"""):
        for c, v in enumerate(row, 1):
            cell = ws2.cell(rr2, c, v); cell.font = BODY
            if c == 6 and v == 0: cell.fill = YEL
        rr2 += 1
    for c in range(1, len(h2) + 1): ws2.column_dimensions[get_column_letter(c)].width = 20
    ws2.column_dimensions["C"].width = 48

    # ---- Sheet 3: ICFR / SOX Control Evidence ----
    ws3 = wb.create_sheet("ICFR_Controls")
    ws3["A1"] = "ICFR/SOX 통제 운영 유효성 (COSO/PCAOB AS 2201)"; ws3["A1"].font = Font(name="Arial", bold=True, size=12)
    h3 = ["control_id", "risk", "frequency", "evidence#", "last", "status"]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(3, c, h); cell.font = HDR; cell.fill = NAVY
    rr3 = 4
    for cid, risk, freq, cnt, last, status in icfr_summary(con):
        cells = [cid, risk, freq, cnt, last, status]
        for c, v in enumerate(cells, 1):
            cell = ws3.cell(rr3, c, v); cell.font = BODY
            if c == 6:
                cell.fill = PatternFill("solid", fgColor="C6EFCE" if status == "effective"
                                        else ("FFC7CE" if status == "no-evidence" else "FFEB9C"))
        rr3 += 1
    for c in range(1, len(h3) + 1): ws3.column_dimensions[get_column_letter(c)].width = 22
    ws3.column_dimensions["B"].width = 30

    # ---- Sheet 4: Variance Bridge (가정 변경 레인 + tie-out + 워터폴 차트) ----
    bridges = list(con.execute("SELECT DISTINCT line_id,scenario_pair FROM variance_lane"))
    if bridges:
        from openpyxl.chart import BarChart, Reference
        ws4 = wb.create_sheet("Variance_Bridge")
        ws4["A1"] = "Variance Bridge — 고정비 (가정 변경 레인 포함)"; ws4["A1"].font = Font(name="Arial", bold=True, size=12)
        row0 = 3
        for line_id, pair in bridges:
            ws4.cell(row0, 1, f"{line_id}  [{pair}]").font = BOLD
            hr = row0 + 1
            for c, h in enumerate(["lane", "amount", "evidence(결정)"], 1):
                cell = ws4.cell(hr, c, h); cell.font = HDR; cell.fill = NAVY
            laned = {ln: (amt, ev) for ln, amt, ev in con.execute(
                "SELECT lane,amount,evidence FROM variance_lane WHERE line_id=? AND scenario_pair=?", (line_id, pair))}
            r = hr + 1; first = r
            for lane in LANES:
                amt, ev = laned.get(lane, (0.0, ""))
                ws4.cell(r, 1, lane).font = (BOLD if lane == "assumption_change" else BODY)
                ac = ws4.cell(r, 2, amt); ac.number_format = KRW; ac.font = BLUE
                if lane == "assumption_change": ac.fill = YEL
                ws4.cell(r, 3, ev).font = BODY
                r += 1
            # tie-out: budget + Σlane == actual (여기선 Σlane 표시 + 0 검증은 콘솔)
            tot = ws4.cell(r, 1, "Σ lanes (=Actual−Budget)"); tot.font = BOLD
            ws4.cell(r, 2, f"=SUM(B{first}:B{r-1})").number_format = KRW
            ws4.cell(r, 2).font = BOLD
            # 워터폴 대용: 레인 막대차트
            chart = BarChart(); chart.title = f"{line_id} variance lanes"; chart.type = "col"; chart.height = 6; chart.width = 14
            data = Reference(ws4, min_col=2, min_row=first, max_row=r - 1)
            cats = Reference(ws4, min_col=1, min_row=first, max_row=r - 1)
            chart.add_data(data, titles_from_data=False); chart.set_categories(cats); chart.legend = None
            ws4.add_chart(chart, f"E{hr}")
            row0 = r + 12
        for c in range(1, 4): ws4.column_dimensions[get_column_letter(c)].width = 26

    # ---- Sheet 5: Ops_SLO + Variance 내러티브 (§6.7 / §7.14) ----
    ws5 = wb.create_sheet("Ops_SLO")
    ws5["A1"] = "에이전트 관측 / SLO (§6.7)"; ws5["A1"].font = Font(name="Arial", bold=True, size=12)
    for c, h in enumerate(["metric", "value", "target", "status"], 1):
        cell = ws5.cell(3, c, h); cell.font = HDR; cell.fill = NAVY
    rr5 = 4
    for metric, value, target, status in compute_slos(con):
        for c, v in enumerate([metric, str(value), target, status], 1):
            cell = ws5.cell(rr5, c, v); cell.font = BODY
            if c == 4: cell.fill = PatternFill("solid", fgColor="C6EFCE" if status == "ok" else ("FFEB9C" if status in ("watch", "info") else "FFC7CE"))
        rr5 += 1
    rr5 += 1
    ws5.cell(rr5, 1, "Variance 내러티브 (그라운딩 검증 통과 문장만, §7.14)").font = BOLD
    for line_id, in con.execute("SELECT DISTINCT line_id FROM variance_lane"):
        nv = narrate_variance(con, line_id)
        if nv["narrative"]:
            rr5 += 1; ws5.cell(rr5, 1, nv["narrative"]).font = BODY; ws5.cell(rr5, 1).alignment = Alignment(wrap_text=True)
    for c in range(1, 5): ws5.column_dimensions[get_column_letter(c)].width = 24
    ws5.column_dimensions["A"].width = 30

    wb.save(path)
    return path


# ============================================================================
# SECTION 13 — 트리거/큐 층 (§4 work_item·router·worker, §6.2 보안, §6.3 outbox)
# ============================================================================
# 결정론은 인입 분류가 아니라 '내가 보낸 발신의 상관키'에 둔다.
# 워커는 inbound를 '데이터(지시 아님)'로 취급하고 카드만 방출(부작용 직접 불가).
def open_request(con, request_id, fcst_line, owner, due_date, now="2026-06-14"):
    token = f"[{request_id}]"
    con.execute("INSERT OR REPLACE INTO request_register VALUES(?,?,?,?,?,?,?)",
                (request_id, fcst_line, owner, due_date, token, "sent", now))
    con.commit(); return token

def enqueue_work_item(con, source, raw_ref, dedup_key, correlation_token=None,
                      payload=None, now="2026-06-14", trust=None):
    """멱등 적재(dedup_key UNIQUE). 상관키 보유=신뢰↑, 미보유=review 전용. trust 명시 시 우선."""
    has_open = bool(correlation_token and con.execute(
        "SELECT 1 FROM request_register WHERE correlation_token=? AND status='sent'", (correlation_token,)).fetchone())
    trust = trust or ("correlated" if has_open else ("internal" if source == "scheduler" else "unsolicited"))
    cur = con.execute("""INSERT OR IGNORE INTO work_item(dedup_key,source,raw_ref,correlation_token,payload,trust,status,created_at)
                         VALUES(?,?,?,?,?,?,?,?)""",
                      (dedup_key, source, raw_ref, correlation_token,
                       json.dumps(payload, ensure_ascii=False) if payload else None, trust, "pending", now))
    con.commit()
    return {"enqueued": cur.rowcount == 1, "dedup_key": dedup_key, "trust": trust}

def route_work_item(con, item) -> Optional[str]:
    """work_item → active playbook. 결정론(상관키/소스) 우선, 없으면 None(→propose_playbook)."""
    wid, source, raw_ref, token, payload, trust, status = item
    if token and con.execute("SELECT 1 FROM request_register WHERE correlation_token=? AND status='sent'", (token,)).fetchone():
        return "inbound_reply_to_request"
    if source == "sharepoint" and raw_ref and "/contracts/" in raw_ref:
        return "contract_ingest"
    if source == "scheduler" and raw_ref and raw_ref.startswith("overdue://"):
        return "overdue_escalation"
    sig = f"{source}:{(raw_ref or '')[:24]}"     # 활성화된 자기확장 플레이북 참조
    act = con.execute("SELECT playbook_id FROM playbook_card WHERE trigger=? AND status='active'", (sig,)).fetchone()
    if act: return act[0]
    return None   # 미매칭 → propose_playbook (자기확장)

def activate_playbook(con, playbook_id, approver, now="2026-06-14"):
    """proposed Playbook 카드 → active(사람 작성·승인 후). 활성화되면 route가 참조해 라우팅."""
    cur = con.execute("UPDATE playbook_card SET status='active' WHERE playbook_id=? AND status='proposed'", (playbook_id,))
    con.commit()
    return {"playbook_id": playbook_id, "status": "active" if cur.rowcount else "not_found", "by": approver}

# 워커 디스패치: 플레이북 → 엔진 호출 → 카드 방출. (실제: 헤드리스 Claude Code 1회)
def run_worker(con, item, playbook, now="2026-06-14"):
    wid, source, raw_ref, token, payload_json, trust, status = item
    payload = json.loads(payload_json) if payload_json else {}
    # 보안(§6.2): 외부 unsolicited는 자동 작업 불가 → 검토 전용(내부 scheduler는 신뢰)
    if trust == "unsolicited":
        con.execute("UPDATE work_item SET status='review' WHERE id=?", (wid,)); con.commit()
        return {"work_item": wid, "action": "review_only(unsolicited, 보안)"}

    if playbook == "inbound_reply_to_request" and payload.get("kind") == "lease_comps":
        # 회신 반환양식(comps)을 파싱해 임차 유불리 분석 실행
        subj = Subject(**payload["subject"])
        comps = [Comp(prov=Provenance(c.pop("uri"), Authority[c.pop("auth")], c.pop("asof")), **c)
                 for c in payload["comps"]]
        res, inp = estimate_market_rent(subj, comps, abs_materiality=payload.get("materiality", 0))
        renewal = res.recommended_value * subj.size_sqm * 12
        proj = project_fcst_lines("lease_favorability", renewal_payment=renewal, renewal_term=5)
        out = submit(con, f"{token} 회신 처리: 임차 유불리", res, inp, projections=proj, fcst_line="HQ_lease_pangyo")
        # 요청 충족 처리
        con.execute("UPDATE request_register SET status='fulfilled' WHERE correlation_token=?", (token,))
        con.execute("UPDATE work_item SET status='done' WHERE id=?", (wid,)); con.commit()
        return {"work_item": wid, "routed": playbook, "result": out}
    if playbook == "contract_ingest" and payload.get("contract"):
        # 계약 인입: 거래처 ER 해소 후 개정 감지(신규/개정·재스케줄·rebuild)
        res_cc = detect_contract_change(con, payload["contract"], now)
        con.execute("UPDATE work_item SET status='done' WHERE id=?", (wid,)); con.commit()
        return {"work_item": wid, "routed": playbook, "result": res_cc}
    if playbook == "overdue_escalation":
        # 독촉: 미충족 요청에 send_reminder Task(발송은 승인→outbox→sink)
        con.execute("""INSERT INTO task_card(analysis_id,task_type,payload,risk_tier,status,requester,approver,created_at)
            VALUES(NULL,'send_reminder',?, 'low','pending_approval','worker',NULL,?)""",
                    (json.dumps(payload, ensure_ascii=False), now))
        con.execute("UPDATE work_item SET status='done' WHERE id=?", (wid,)); con.commit()
        return {"work_item": wid, "routed": playbook, "result": "send_reminder Task(승인 대기)"}
    if payload.get("kind") == "note":
        # 노트 파생(사람 확인됨): fcst 가정 검토 Task + 요청 충족
        con.execute("""INSERT INTO task_card(analysis_id,task_type,payload,risk_tier,status,requester,approver,created_at)
            VALUES(NULL,'review_note',?, 'low','pending_approval','worker',NULL,?)""",
                    (json.dumps({"fcst_line": payload.get("fcst_line"), "topic": payload.get("topic")}, ensure_ascii=False), now))
        con.execute("UPDATE request_register SET status='fulfilled' WHERE correlation_token=?", (token,))
        con.execute("UPDATE work_item SET status='done' WHERE id=?", (wid,)); con.commit()
        return {"work_item": wid, "routed": playbook, "result": "note → review_note Task"}
    con.execute("UPDATE work_item SET status='done' WHERE id=?", (wid,)); con.commit()
    return {"work_item": wid, "routed": playbook, "result": "handled"}

def drain(con, now="2026-06-14", max_attempts=3):
    """스케줄 드레인: pending work_item을 라우팅·워커 디스패치. (Task Scheduler 진입점)
    결함격리: 한 건 실패가 배치를 중단시키지 않음 — attempts 누적, max 도달 시 dead-letter."""
    results = []
    scan_overdue_requests(con, now)   # 미충족 요청 → overdue work_item (이번 패스서 처리)
    for item in con.execute("""SELECT id,source,raw_ref,correlation_token,payload,trust,status
                               FROM work_item WHERE status='pending' ORDER BY id""").fetchall():
        wid = item[0]; pb = route_work_item(con, item)
        if pb is None:
            sig = f"{item[1]}:{(item[2] or '')[:24]}"   # 자기확장: 미매칭 → proposed Playbook 카드(중복 무시)
            con.execute("INSERT OR IGNORE INTO playbook_card(playbook_id,trigger,status,proposed_for,created_at) VALUES(?,?,?,?,?)",
                        (f"PB-{sig}", sig, "proposed", item[2], now))
            con.execute("UPDATE work_item SET status='awaiting_playbook' WHERE id=?", (wid,)); con.commit()
            results.append({"work_item": wid, "action": "propose_playbook(미매칭→proposed 카드)"})
            st = "no_playbook"
        else:
            try:
                results.append(run_worker(con, item, pb, now)); st = "completed"
            except Exception as e:   # 결함격리 + dead-letter
                att = (con.execute("SELECT attempts FROM work_item WHERE id=?", (wid,)).fetchone()[0] or 0) + 1
                new = "dead" if att >= max_attempts else "pending"
                con.execute("UPDATE work_item SET status=?, attempts=? WHERE id=?", (new, att, wid)); con.commit()
                results.append({"work_item": wid, "error": str(e)[:80], "attempts": att, "status": new})
                st = "failed"
        con.execute("INSERT OR REPLACE INTO ops_run VALUES(?,?,?,?,?,?,?,?)",
                    (f"run-wi{wid}-a{con.execute('SELECT attempts FROM work_item WHERE id=?', (wid,)).fetchone()[0]}",
                     "worker", now, now, st, 1500, 0.02, json.dumps({"playbook": pb, "trust": item[5]}, ensure_ascii=False)))
    rb = process_rebuild_requests(con, now)   # 개정으로 등록된 rebuild 소비 → 무효화+재예측 Task
    if rb: results.append({"rebuilds": rb})
    con.commit()
    return results

# --- Transactional outbox executor (§6.3): exactly-once 부작용 + ICFR 완성 ---
def enqueue_outbox(con, task_id, side_effect, idem_key, now="2026-06-14"):
    con.execute("""INSERT OR IGNORE INTO outbox(task_id,side_effect,idempotency_key,status,created_at)
                   VALUES(?,?,?,?,?)""", (task_id, side_effect, idem_key, "pending", now))
    con.commit()

def process_outbox(con, now="2026-06-14", sender="log"):
    """outbox 디스패처: pending을 1회 실행. idempotency_key로 중복 차단(크래시 복구).
    부작용은 sender(기본 log_sender→sent_log 기록)로 실행; 외부(COM/GL)는 SENDERS 교체."""
    done = []; send = SENDERS.get(sender, log_sender)
    for oid, task_id, side_effect, idem, st in con.execute(
            "SELECT id,task_id,side_effect,idempotency_key,status FROM outbox WHERE status='pending' ORDER BY id"):
        external_ref = send(con, task_id, idem, side_effect, now)   # 실제 실행(sink)
        aid = con.execute("SELECT analysis_id FROM task_card WHERE id=?", (task_id,)).fetchone()[0]
        con.execute("UPDATE task_card SET status='done', executed_result=? WHERE id=?",
                    (json.dumps({"external_ref": external_ref}, ensure_ascii=False), task_id))
        con.execute("UPDATE control_evidence SET approval_status='approved' WHERE analysis_id=?", (aid,))
        con.execute("UPDATE decision_analysis SET status='applied', linked_assumption_card=? WHERE id=? AND status='review'",
                    (f"assumption::{side_effect}", aid))
        con.execute("UPDATE outbox SET status='confirmed', external_ref=? WHERE id=?", (external_ref, oid))
        con.commit()
        done.append({"outbox": oid, "task": task_id, "external_ref": external_ref})
    return done


# ============================================================================
# SECTION 14 — Variance Bridge (§3 + "가정 변경" 레인; 결정→숫자 추적)
# ============================================================================
# 고정비 variance 레인: contract_change / new_asset / one_time / indexation /
#                       assumption_change(결정 귀속) / residual(잔차).
LANES = ["contract_change", "new_asset", "one_time", "indexation", "assumption_change", "residual"]

def add_variance_lane(con, line_id, lane, amount, evidence="", scenario_pair="Budget->Actual", now="2026-06-14"):
    con.execute("INSERT INTO variance_lane(line_id,scenario_pair,lane,amount,evidence,created_at) VALUES(?,?,?,?,?,?)",
                (line_id, scenario_pair, lane, amount, evidence, now)); con.commit()

def assumption_change_from_decisions(con, line_id, budget=None):
    """이 라인의 결정들이 모델링한 1차년 P&L − 예산 = 가정변경(re-forecast) 귀속액 + 근거 라벨.
    근거: 유연예산/재예측 분산분해 — 가정변경분 = 모델 예측 − 계획(플렉서블 버짓 variance)."""
    ev = []; proj_y1 = 0.0; has_proj = False
    for aid, domain in con.execute(
            "SELECT id,domain FROM decision_analysis WHERE fcst_line=? AND status IN ('applied','review')", (line_id,)):
        ev.append(f"D-{aid}:{domain}")
        for (v,) in con.execute("SELECT value FROM fcst_line_projection WHERE analysis_id=? AND period=1", (aid,)):
            proj_y1 += v; has_proj = True
    amount = round(proj_y1 - budget, 2) if (has_proj and budget is not None) else 0.0
    return amount, ev

def build_variance_bridge(con, line_id, budget, actual, lanes: dict, now="2026-06-14"):
    """레인 적재 + tie-out(budget + Σlane == actual). residual은 잔차로 자동.
    assumption_change는 결정 투영에서 산출(전달값이 없을 때); residual이 나머지를 흡수."""
    named = {k: v for k, v in lanes.items() if k != "residual"}
    if "assumption_change" not in named:
        amt_ac, _ = assumption_change_from_decisions(con, line_id, budget)
        if amt_ac: named["assumption_change"] = amt_ac     # 결정에서 산출된 귀속액
    residual = actual - budget - sum(named.values())
    for lane in LANES:
        amt = named.get(lane, 0.0) if lane != "residual" else residual
        ev = ""
        if lane == "assumption_change":
            _, decev = assumption_change_from_decisions(con, line_id, budget)
            ev = ",".join(decev)
        add_variance_lane(con, line_id, lane, amt, ev, now=now)
    tie = (budget + sum(named.values()) + residual) - actual   # ==0
    return {"line_id": line_id, "budget": budget, "actual": actual,
            "lanes": {**named, "residual": residual}, "tie_out": round(tie, 2)}


# ============================================================================
# SECTION 16 — Reference Data 레이어 (데이터 획득 스펙: 외부지표→reference_data)
# ============================================================================
# 폐쇄망 원칙: 외부 직접 호출 금지. 어댑터가 날짜 스냅샷을 SharePoint에 착지 →
# 여기서 reference_data(SCD2)로 적재. 엔진은 리터럴 대신 이 테이블 조회.
import hashlib

def ingest_snapshot(con, table, rows, keys, ref_set, source, url, license, version, now="2026-06-14"):
    """SCD2 적재(멱등): 동일 키 현행 행 close 후 신규 open + 스냅샷 메타."""
    sha = hashlib.sha256(json.dumps(rows, sort_keys=True, ensure_ascii=False).encode()).hexdigest()[:16]
    exists = con.execute("SELECT 1 FROM ref_snapshot WHERE ref_set=? AND sha256=?", (ref_set, sha)).fetchone()
    con.execute("INSERT OR IGNORE INTO ref_snapshot VALUES(?,?,?,?,?,?,?,?)",
                (ref_set, now, source, url, license, version, sha, now))
    if exists:   # 동일 내용 재적재 → SCD2 행 무변경(0-duration 행/중복 방지, 멱등)
        con.commit()
        return {"ref_set": ref_set, "version": version, "sha256": sha, "rows": 0, "idempotent": True}
    inserted = 0
    for r in rows:
        where = " AND ".join(f"{k}=?" for k in keys)
        con.execute(f"UPDATE {table} SET valid_to=? WHERE {where} AND valid_to IS NULL",
                    (now, *[r[k] for k in keys]))
        cols = list(r) + ["source_version", "valid_from", "valid_to"]
        con.execute(f"INSERT INTO {table}({','.join(cols)}) VALUES({','.join(['?']*len(cols))})",
                    (*r.values(), version, now, None))
        inserted += 1
    con.commit()
    return {"ref_set": ref_set, "version": version, "sha256": sha, "rows": inserted}

# --- 모의 외부 스냅샷(실제: 어댑터가 data.go.kr/ECOS OpenAPI에서 받아 SharePoint 착지) ---
def seed_reb_snapshot(con, version, now="2026-06-14"):
    """한국부동산원 상업용부동산 임대동향조사(오피스) 지역×등급×분기 (단위 원/㎡/월)."""
    rows = [
        {"region": "판교", "property_type": "office", "grade": 4, "period": "2025Q4", "rent_per_sqm": 19800, "rent_index": 101.2, "vacancy": 4.1},
        {"region": "판교", "property_type": "office", "grade": 4, "period": "2026Q1", "rent_per_sqm": 20050, "rent_index": 102.5, "vacancy": 3.8},
        {"region": "이천", "property_type": "office", "grade": 3, "period": "2026Q1", "rent_per_sqm": 8200, "rent_index": 99.8, "vacancy": 6.2},
    ]
    return ingest_snapshot(con, "regional_rent_benchmark", rows, ["region", "property_type", "grade", "period"],
                           "REB_office_rent", "한국부동산원 상업용부동산 임대동향조사",
                           "data.go.kr/data/15069787", "국가통계(출처명시)", version, now)

def seed_ecos_ibr(con, version, now="2026-06-14"):
    """ECOS 회사채(AA-,3년)·국고채 벤치마크 + 내부 스프레드 → IBR 매트릭스."""
    rows = [
        {"currency": "KRW", "term_band": "3-5y", "security": "secured", "ibr": 0.052, "benchmark": "회사채AA-3y", "spread": 0.004},
        {"currency": "KRW", "term_band": "3-5y", "security": "unsecured", "ibr": 0.061, "benchmark": "회사채AA-3y", "spread": 0.013},
    ]
    return ingest_snapshot(con, "ibr_matrix", rows, ["currency", "term_band", "security"],
                           "ECOS_ibr", "한국은행 ECOS 시장금리", "ecos.bok.or.kr/api", "출처명시 무료", version, now)

# --- 엔진 리팩터: 리터럴 → reference_data 조회 ---
def regional_params(con, region, grade, asof_period):
    """REB 지표에서 시장 수준·월추세 유도(literal market_trend_pm 대체)."""
    cur = con.execute("""SELECT rent_per_sqm,rent_index,period FROM regional_rent_benchmark
        WHERE region=? AND grade=? AND period<=? AND valid_to IS NULL ORDER BY period DESC LIMIT 2""",
        (region, grade, asof_period)).fetchall()
    if not cur: return None
    trend_pm = ((cur[0][1] / cur[1][1]) ** (1 / 3) - 1) if len(cur) == 2 and cur[1][1] else 0.0  # 분기지수→월
    return {"market_rent_anchor": cur[0][0], "market_trend_pm": round(trend_pm, 5), "period": cur[0][2],
            "source": "REB"}

def get_ibr(con, currency="KRW", term_band="3-5y", security="secured"):
    row = con.execute("""SELECT ibr,benchmark,spread,source_version FROM ibr_matrix
        WHERE currency=? AND term_band=? AND security=? AND valid_to IS NULL""",
        (currency, term_band, security)).fetchone()
    return {"ibr": row[0], "benchmark": row[1], "spread": row[2], "version": row[3]} if row else None

def ref_freshness(con, ref_set, sla_days, asof="2026-06-14"):
    """데이터 계약 freshness: 최신 스냅샷이 SLA 초과면 stale."""
    from datetime import date
    row = con.execute("SELECT MAX(snapshot_date) FROM ref_snapshot WHERE ref_set=?", (ref_set,)).fetchone()
    if not row or not row[0]: return {"ref_set": ref_set, "status": "missing"}
    age = (date.fromisoformat(asof) - date.fromisoformat(row[0])).days
    return {"ref_set": ref_set, "age_days": age, "status": "stale" if age > sla_days else "fresh"}


# ============================================================================
# SECTION 17 — 시간 기반 신뢰 감쇠 (§7.7: 노후 가정 → stale → 재확인 트리거)
# ============================================================================
# 근거 종류별 반감기(월): 계약성 근거는 느리게, 시장추정·구두는 빠르게.
HALF_LIFE_MONTHS = {"contractual": 36.0, "projection": 12.0, "market_estimate": 6.0, "verbal": 2.0}

def _months_between(a, b):
    from datetime import date
    da, db = date.fromisoformat(a), date.fromisoformat(b)
    return (db.year - da.year) * 12 + (db.month - da.month) + (db.day - da.day) / 30.0

def apply_confidence_decay(con, asof, stale_threshold=0.4, now=None):
    """경과로 신뢰도 감쇠. 임계 하회 시 status=stale + 재확인 Task(자료요청) 자동 제안."""
    now = now or asof
    out = []
    for aid, ek, conf, created, status, fcst_line in con.execute(
            """SELECT id,evidence_kind,confidence,created_at,status,fcst_line FROM decision_analysis
               WHERE status IN ('applied','review','provisional')"""):
        hl = HALF_LIFE_MONTHS.get(ek or "market_estimate", 6.0)
        age = max(0.0, _months_between(created, asof))
        decayed = conf * (0.5 ** (age / hl))
        if decayed < stale_threshold:
            con.execute("UPDATE decision_analysis SET status='stale' WHERE id=?", (aid,))
            # 재확인 Task(사람 승인) + 자료요청 레지스터 오픈
            con.execute("""INSERT INTO task_card(analysis_id,task_type,payload,risk_tier,status,requester,approver,created_at)
                           VALUES(?,?,?,?,?,?,?,?)""",
                        (aid, "propose_data_request",
                         json.dumps({"reason": "confidence decay", "decayed": round(decayed, 3),
                                     "fcst_line": fcst_line, "evidence_kind": ek}, ensure_ascii=False),
                         "low", "pending_approval", "system", None, now))
            rid = f"REQ-RC-{aid}"
            con.execute("INSERT OR REPLACE INTO request_register VALUES(?,?,?,?,?,?,?)",
                        (rid, fcst_line, "fixedcost-fpna", asof, f"[{rid}]", "sent", now))
            out.append({"analysis_id": aid, "evidence_kind": ek, "age_months": round(age, 1),
                        "decayed_conf": round(decayed, 3), "action": "stale→재확인 요청"})
    con.commit()
    return out


# ============================================================================
# SECTION 18 — 콜드스타트 / 역사 백필 (§7.11: 1일차 부트스트랩)
# ============================================================================
def bootstrap_from_history(con, contracts, gl_baseline, now="2026-06-14"):
    """기존 계약·과거 GL로 초기 원장·기준선 부트스트랩. 카드는 provisional(일괄 검토 대상)."""
    created = []
    for c in contracts:
        # 기존 계약 → provisional assumption(출처=기존 문서, 약한 확신). 일반 흐름 배치모드.
        res = AnalysisResult(domain=c["domain"], recommended_value=c["value"],
                             recommendation=f"기존 계약 인입(검토 전 provisional): {c['name']}",
                             confidence=0.55, model_version="bootstrap/1.0.0",
                             materiality_band="n/a", grounded=True,
                             key_assumptions={"bootstrap": True, "doc": c["doc"]})
        inp = [Input(c["name"], c["value"], "KRW",
                     Provenance(c["doc"], Authority.SIGNED_CONTRACT, c.get("asof", "2024-01-01")))]
        cur = con.execute("""INSERT INTO decision_analysis(question,domain,fcst_line,evidence_kind,inputs,model,model_version,
                result,confidence,conf_interval,sensitivity,materiality_band,grounded,status,linked_assumption_card,next_review,created_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (f"부트스트랩: {c['name']}", c["domain"], c["fcst_line"], "contractual",
             json.dumps([{"name": i.name, "value": i.value, "unit": i.unit, "source_uri": i.prov.source_uri,
                          "source_authority": int(i.prov.authority), "asof": i.prov.asof} for i in inp], ensure_ascii=False),
             "bootstrap", "bootstrap/1.0.0", json.dumps({"recommendation": res.recommendation, "value": c["value"]}, ensure_ascii=False),
             0.55, "[0,0]", "{}", "n/a", 1, "provisional", None, now, c.get("asof", "2024-01-01")))
        created.append({"fcst_line": c["fcst_line"], "analysis_id": cur.lastrowid, "status": "provisional"})
    # 과거 GL 기준선(variance baseline) 적재
    for line, periods in gl_baseline.items():
        for p, v in periods.items():
            con.execute("INSERT INTO fcst_line_projection(analysis_id,line_type,period,value,note) VALUES(NULL,?,?,?,?)",
                        (f"baseline:{line}", p, v, "콜드스타트 과거 GL 기준선"))
    con.commit()
    return created

def bulk_confirm_provisional(con, approver, now="2026-06-14"):
    """부트스트랩 provisional 카드 일괄 검토 승인 → confirmed(applied)."""
    n = con.execute("SELECT COUNT(*) FROM decision_analysis WHERE status='provisional'").fetchone()[0]
    con.execute("UPDATE decision_analysis SET status='applied', linked_assumption_card='assumption::bootstrap' WHERE status='provisional'")
    con.commit()
    return {"confirmed": n, "by": approver}


# ============================================================================
# SECTION 20 — 거래처 Entity Resolution (§7.9: Fellegi-Sunter, stdlib)
# ============================================================================
# Splink는 DuckDB 의존이라 미설치 — Fellegi-Sunter 알고리즘만 stdlib 재현.
# 명칭 변형(㈜/주식회사/영문/오타)을 사업자번호·이름·주소 비교로 클러스터링.
import difflib

_LEGAL = ["주식회사", "(주)", "㈜", "co.,ltd", "co., ltd", "co.ltd", ", ltd", " ltd", " inc", " corp"]
_COMMON = {"물류", "코리아", "korea", "logistics", "company", "co"}   # TF: 흔한 토큰 다운웨이트

def normalize_name(s):
    t = s.lower().replace(".", "").replace(",", " ")
    for suf in _LEGAL: t = t.replace(suf, " ")
    return " ".join(t.split())

def _name_level(a, b):
    na, nb = normalize_name(a), normalize_name(b)
    sim = difflib.SequenceMatcher(None, na, nb).ratio()
    # TF 보정: 공통 토큰만으로 겹치면 강한 일치로 보지 않음
    shared = (set(na.split()) & set(nb.split()))
    only_common = shared and shared <= _COMMON
    if sim >= 0.92 and not only_common: return "exact", sim
    if sim >= 0.70: return "partial", sim
    return "low", sim

# 레벨별 매치 가중치(log2(m/u) 유도값). 양수=일치 증거, 음수=불일치.
_W_BIZ = {"equal": 13.0, "diff": -7.0, "na": 0.0}
_W_NAME = {"exact": 5.5, "partial": 2.0, "low": -3.0}
_W_ADDR = {"agree": 2.0, "disagree": -1.0, "na": 0.0}
_UPPER, _LOWER = 6.0, 0.0      # >=upper: auto-match, [lower,upper): review, <lower: 비매치

def _pair_score(r1, r2, w_biz_agree=None):
    s = 0.0
    if r1.get("biz_no") and r2.get("biz_no"):
        s += (w_biz_agree if w_biz_agree is not None else _W_BIZ["equal"]) if r1["biz_no"] == r2["biz_no"] else _W_BIZ["diff"]
    lvl, _ = _name_level(r1["name"], r2["name"]); s += _W_NAME[lvl]
    if r1.get("addr") and r2.get("addr"):
        ta, tb = set(normalize_name(r1["addr"]).split()), set(normalize_name(r2["addr"]).split())
        ov = len(ta & tb) / max(1, len(ta | tb))
        s += _W_ADDR["agree"] if ov >= 0.4 else _W_ADDR["disagree"]
    return s

def resolve_entities(con, records, now="2026-06-14", learn=True):
    """records: [{id,name,biz_no,addr}]. union-find로 auto-match 클러스터링 → canonical.
    스케일(n≥30): EM으로 m 학습 + **값별 TF 보정**(희소 사업자번호=강증거) — 일치값 x의
    가중치=log2(m/u_x), u_x=빈도(x)/N. 근거: Splink term-frequency, Xu·Li·Grannis(2021)."""
    n = len(records); m_learn = None; u_field = None; counts = {}
    if learn and n >= 30:
        p = train_er_em(records, "biz_no")
        if p["n_pairs"]:
            m_learn, u_field = p["m"], p["u"]
            for r in records:
                b = r.get("biz_no")
                if b: counts[b] = counts.get(b, 0) + 1
    parent = {r["id"]: r["id"] for r in records}
    def find(x):
        while parent[x] != x: parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a, b): parent[find(a)] = find(b)
    review = []
    for i in range(len(records)):
        for j in range(i + 1, len(records)):
            wb = None
            if m_learn is not None and records[i].get("biz_no") and records[i]["biz_no"] == records[j].get("biz_no"):
                u_x = max(counts.get(records[i]["biz_no"], 1) / n, 1.0 / n)
                # 희소값만 강화, 고유 식별자는 기저(u_field) 미만 약화 방지 → min
                wb = math.log2(m_learn / min(u_field, u_x))
            sc = _pair_score(records[i], records[j], wb)
            if sc >= _UPPER: union(records[i]["id"], records[j]["id"])
            elif sc >= _LOWER: review.append((records[i]["id"], records[j]["id"], round(sc, 1)))
    # canonical = 클러스터 내 최소 id
    clusters = {}
    for r in records: clusters.setdefault(find(r["id"]), []).append(r)
    out = []
    for root, members in clusters.items():
        cvid = f"V-{min(m['id'] for m in members)}"
        for m in members:
            status = "auto" if len(members) > 1 else "singleton"
            con.execute("INSERT INTO vendor_cluster VALUES(?,?,?,?,?,?,?)",
                        (m["name"], m.get("biz_no"), cvid, 0.0, "fellegi_sunter", status, now))
            out.append({"raw_name": m["name"], "canonical": cvid})
    con.commit()
    return {"clusters": {f"V-{min(m['id'] for m in mem)}": [m["name"] for m in mem] for mem in clusters.values()},
            "review_pairs": review}

def resolve_counterparty(con, name, biz_no=None):
    """이미 클러스터된 vendor_cluster에서 canonical 조회(사업자번호 우선)."""
    if biz_no:
        row = con.execute("SELECT canonical_vendor_id FROM vendor_cluster WHERE biz_no=? LIMIT 1", (biz_no,)).fetchone()
        if row: return row[0]
    row = con.execute("SELECT canonical_vendor_id FROM vendor_cluster WHERE raw_name=? LIMIT 1", (name,)).fetchone()
    return row[0] if row else None


# ============================================================================
# SECTION 21 — 계약 개정 감지 (§7.8: 부속합의=기존 계약 새 버전, 스케줄 재도출)
# ============================================================================
def detect_contract_change(con, incoming, now="2026-06-14"):
    """incoming 계약을 기존 마스터와 매칭(ER 거래처+자산). 매칭=개정(supersede+재스케줄+rebuild), 아니면 신규."""
    cvid = resolve_counterparty(con, incoming["counterparty_name"], incoming.get("biz_no")) or incoming["counterparty_name"]
    existing = con.execute("""SELECT contract_id,version,amendment_seq FROM contract_master
        WHERE counterparty=? AND asset_or_property=? AND status='active'""",
        (cvid, incoming["asset_or_property"])).fetchone()
    if existing:
        old_id, ver, seq = existing
        new_id = f"{incoming['contract_no']}.v{ver + 1}"
        con.execute("UPDATE contract_master SET status='superseded' WHERE contract_id=?", (old_id,))
        con.execute("""INSERT INTO contract_master VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (new_id, cvid, incoming["asset_or_property"], incoming["contract_no"], ver + 1, seq + 1, old_id,
                     incoming["monthly_amount"], incoming["term_months"], incoming.get("escalation", 0.0),
                     incoming.get("start_date", now), incoming.get("end_date"), "active", incoming["source_doc"],
                     incoming.get("fcst_line"), now))
        # 스케줄 재도출(리스) + 영향 자산 retraction 전파(§6.6 rebuild)
        _ibr = get_ibr(con, "KRW", "3-5y", "secured")
        ibr = _ibr["ibr"] if _ibr else 0.052
        sched = lease_schedule(LeasePlan(incoming["monthly_amount"] * 12, incoming["term_months"] // 12 or 1, ibr, escalation=incoming.get("escalation", 0.0)))
        con.execute("INSERT INTO rebuild_request(trigger,affected,status,created_at) VALUES(?,?,?,?)",
                    (f"amendment:{new_id}", incoming.get("fcst_line"), "pending", now))
        con.commit()
        # 리스 변경=재측정(IFRS 16) 회계 이벤트 — 통지 토글 ON일 때만 회계팀 통지
        if is_accounting_notify_on("lease_remeasurement"):
            con.execute("""INSERT INTO task_card(analysis_id,task_type,payload,risk_tier,status,requester,approver,created_at)
                VALUES(NULL,'notify_accounting',?, 'low','pending_approval','system',NULL,?)""",
                (json.dumps({"event": "lease_remeasurement", "fcst_line": incoming.get("fcst_line"),
                             "summary": f"계약 개정 {new_id}", "standard": STD_REF["lease_remeasurement"]}, ensure_ascii=False), now))
            con.commit()
        return {"type": "amendment", "new_contract": new_id, "supersedes": old_id,
                "amendment_seq": seq + 1, "counterparty": cvid, "rescheduled_periods": len(sched), "rebuild": "registered"}
    new_id = f"{incoming['contract_no']}.v1"
    con.execute("""INSERT INTO contract_master VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (new_id, cvid, incoming["asset_or_property"], incoming["contract_no"], 1, 0, None,
                 incoming["monthly_amount"], incoming["term_months"], incoming.get("escalation", 0.0),
                 incoming.get("start_date", now), incoming.get("end_date"), "active", incoming["source_doc"],
                 incoming.get("fcst_line"), now))
    con.commit()
    return {"type": "new", "contract": new_id, "counterparty": cvid}

def process_rebuild_requests(con, now="2026-06-14"):
    """§6.6 rebuild: 계약 개정으로 등록된 rebuild_request 소비 — 영향 라인의 적용 결정을 stale로
    무효화하고 propose_reforecast Task(사람 승인) 방출. 개정→하류 투영 재산출 seam 연결."""
    out = []
    for rid, trigger, affected, status, created in con.execute(
            "SELECT id,trigger,affected,status,created_at FROM rebuild_request WHERE status='pending'"):
        n_inv = 0
        if affected:
            ids = [r[0] for r in con.execute("SELECT id FROM decision_analysis WHERE fcst_line=? AND status IN('applied','review')", (affected,))]
            for aid in ids:
                con.execute("UPDATE decision_analysis SET status='stale' WHERE id=?", (aid,))
                con.execute("""INSERT INTO task_card(analysis_id,task_type,payload,risk_tier,status,requester,approver,created_at)
                    VALUES(?,'propose_reforecast',?, 'med','pending_approval','system',NULL,?)""",
                    (aid, json.dumps({"reason": trigger, "fcst_line": affected}, ensure_ascii=False), now))
            n_inv = len(ids)
        con.execute("UPDATE rebuild_request SET status='processed' WHERE id=?", (rid,))
        out.append({"rebuild": rid, "trigger": trigger, "fcst_line": affected,
                    "invalidated_decisions": n_inv, "action": "stale+propose_reforecast Task"})
    con.commit()
    return out

def scan_overdue_requests(con, asof="2026-06-14"):
    """미충족 자료요청(due_date 경과·status='sent') → overdue_escalation work_item 적재(드레인서 처리)."""
    out = []
    for rid, fcst_line, due in con.execute(
            "SELECT request_id,fcst_line,due_date FROM request_register WHERE status='sent'"):
        if due and due < asof:
            r = enqueue_work_item(con, "scheduler", f"overdue://{rid}", f"overdue:{rid}",
                                  payload={"kind": "overdue", "request_id": rid, "fcst_line": fcst_line}, now=asof)
            out.append({"request": rid, "due": due, "enqueued": r["enqueued"]})
    return out


# ============================================================================
# SECTION 22 — 민감도 / What-if 엔진 (§7.3: 드라이버 분기, 원장 불변·발송 없음)
# ============================================================================
def lease_sensitivity(con, base_ref, base_payment, base_term_yr, base_escalation, ibr,
                      escalations, renew_options, now="2026-06-14"):
    """임차 갱신 민감도: 인상률 × 갱신여부 → 기간 총 P&L + base 대비 델타."""
    def pnl_total(esc, term):
        if term == 0: return 0.0   # 비갱신(이전/철수) 단순화
        return sum(r[6] for r in lease_schedule(LeasePlan(base_payment, term, ibr, escalation=esc)))
    base_total = pnl_total(base_escalation, base_term_yr)
    cases = []
    for esc in escalations:
        for renew in renew_options:
            term = base_term_yr if renew else 0
            total = pnl_total(esc, term); delta = total - base_total
            cid = f"{base_ref}:esc{int(esc*1000)}_renew{int(renew)}"
            con.execute("INSERT INTO sensitivity_case VALUES(?,?,?,?,?,?,?,?,?)",
                        (cid, base_ref, "lease", json.dumps({"escalation": esc, "renew": renew}),
                         "lease_pnl_total", round(total, 0), round(delta, 0), 0, now))
            cases.append({"esc": esc, "renew": renew, "total": round(total, 0), "delta": round(delta, 0)})
    con.commit()
    return {"base_total": round(base_total, 0), "cases": cases}

def buy_vs_lease_sensitivity(con, base_ref, own: OwnPlan, lease: LeasePlan, hurdles, horizon, now="2026-06-14"):
    """할인율 민감도: 권고(buy/lease)가 뒤집히는지(의사결정 안정성)."""
    base = None; cases = []
    for h in hurdles:
        diff = npv(h, _own_cf(own, horizon)) - npv(h, _lease_cf(lease, own.tax_rate, horizon))
        rec = "buy" if diff > 0 else "lease"
        if base is None: base = rec
        flips = int(rec != base)
        cid = f"{base_ref}:hurdle{int(h*1000)}"
        con.execute("INSERT INTO sensitivity_case VALUES(?,?,?,?,?,?,?,?,?)",
                    (cid, base_ref, "buy_vs_lease", json.dumps({"hurdle": h}), "own_minus_lease_npv",
                     round(diff, 0), 0, flips, now))
        cases.append({"hurdle": h, "rec": rec, "own_minus_lease": round(diff, 0), "flips": bool(flips)})
    con.commit()
    return {"base_rec": base, "cases": cases}


# ============================================================================
# SECTION 23 — 지속 eval / 회귀 (§6.9: 라우터·게이트 보정·그라운딩)
# ============================================================================
# 코드 테스트와 별개. 모델/프롬프트/플레이북 변경은 이 회귀를 통과해야 배포(임계 하회 시 차단).
def eval_router(con, now="2026-06-14"):
    """라우터 매칭 정확도. 골든: (work_item, 기대 플레이북)."""
    # 골든용 open request 시드
    con.execute("INSERT OR REPLACE INTO request_register VALUES('REQ-EV','EV_line','o','2026-12-31','[REQ-EV]','sent',?)", (now,))
    golden = [
        (("X", "outlook", "msg", "[REQ-EV]", None, "correlated", "pending"), "inbound_reply_to_request"),
        (("X", "sharepoint", "/contracts/x.pdf", None, None, "unsolicited", "pending"), "contract_ingest"),
        (("X", "outlook", "msg://misc", None, None, "unsolicited", "pending"), None),
    ]
    ok = sum(1 for item, exp in golden if route_work_item(con, item) == exp)  # item = 7-튜플(id,source,raw_ref,token,payload,trust,status)
    acc = ok / len(golden)
    return _record_eval(con, "router_match", "accuracy", acc, 0.95, now)

def eval_gate_calibration(con, now="2026-06-14"):
    """게이트 보정: false-apply(틀린데 자동) / false-escalate(맞는데 검토) 율."""
    golden = [  # (grounded, conf, materiality, tier, should_auto)
        (True, 0.80, "below", "med", True), (True, 0.80, "material", "med", False),
        (False, 0.90, "below", "med", False), (True, 0.50, "below", "med", False),
        (True, 0.90, "below", "high", False), (True, 0.95, "near", "med", True)]
    fa = sum(1 for g, c, m, t, lab in golden if gate_decision(g, c, m, t) and not lab)
    fe = sum(1 for g, c, m, t, lab in golden if (not gate_decision(g, c, m, t)) and lab)
    false_rate = (fa + fe) / len(golden)
    return _record_eval(con, "gate_calibration", "false_rate", false_rate, 0.05, now, le=True,
                        detail=f"false_apply={fa},false_escalate={fe}")

def eval_grounding(con, now="2026-06-14"):
    """그라운딩 판정 정확도: verify_claim(수치+어휘)을 라벨 세트(정상/환각)에 적용해 측정."""
    ev = "임차료가 6,600만원 감소했고 주 요인은 재협상"
    cases = [(ev, "6600만원 감소", True), (ev, "재협상이 주 요인", True),
             (ev, "신규 차량이 원인", False), (ev, "임차료 9999만원 증가", False)]
    acc = sum(1 for e, claim, exp in cases if verify_claim(e, claim)["grounded"] == exp) / len(cases)
    return _record_eval(con, "grounding", "accuracy", acc, 0.90, now)

def _record_eval(con, suite, metric, value, threshold, now, le=False, detail=""):
    passed = (value <= threshold) if le else (value >= threshold)
    con.execute("INSERT INTO eval_run(suite,metric,value,threshold,passed,detail,run_at) VALUES(?,?,?,?,?,?,?)",
                (suite, metric, round(value, 4), threshold, int(passed), detail, now)); con.commit()
    return {"suite": suite, "metric": metric, "value": round(value, 4), "threshold": threshold,
            "cmp": "≤" if le else "≥", "passed": bool(passed), "detail": detail}

def eval_deploy_gate(con, now="2026-06-14", cost_budget=1.0):
    """전 스위트 + 비용 회귀 → 하나라도 실패 시 배포 차단."""
    suites = [eval_router(con, now), eval_gate_calibration(con, now), eval_grounding(con, now)]
    cost = eval_cost_regression(con, cost_budget)
    blocked = [s["suite"] for s in suites if not s["passed"]] + (["cost"] if cost["regressed"] else [])
    return {"suites": suites, "cost": cost, "deploy": "BLOCKED" if blocked else "PASS", "failing": blocked}


# ============================================================================
# SECTION 24 — Variance 내러티브 생성 (§7.14: 그라운딩 — 환각 차단)
# ============================================================================
def narrate_variance(con, line_id, scenario_pair="Budget->Actual"):
    """variance bridge에서 CEO 코멘터리 생성. 각 문장은 레인 수치로 cite-back 검증, 미지지 차단."""
    lanes = {ln: (amt, ev) for ln, amt, ev in con.execute(
        "SELECT lane,amount,evidence FROM variance_lane WHERE line_id=? AND scenario_pair=?", (line_id, scenario_pair))}
    if not lanes: return {"narrative": "", "all_grounded": True, "blocked": []}
    total = sum(a for a, _ in lanes.values())
    sents = [{"text": f"{line_id}: 예산 대비 ₩{abs(total):,.0f} {'감소' if total < 0 else '증가'}.",
              "claim": total, "ref": "Σlanes"}]
    for lane, (amt, ev) in sorted(lanes.items(), key=lambda x: -abs(x[1][0])):
        if abs(amt) < 1: continue
        sents.append({"text": f"· {lane} ₩{amt:+,.0f}{(' [근거 ' + ev + ']') if ev else ''}.",
                      "claim": amt, "ref": lane, "lane": lane})
    # 그라운딩: 각 문장의 수치가 데이터와 일치하는가
    for s in sents:
        s["grounded"] = (abs((lanes[s["lane"]][0] if "lane" in s else total) - s["claim"]) < 1)
    final = [s for s in sents if s["grounded"]]
    blocked = [s for s in sents if not s["grounded"]]
    return {"narrative": " ".join(s["text"] for s in final),
            "all_grounded": all(s["grounded"] for s in sents), "blocked": [b["text"] for b in blocked]}

def verify_narrative_claim(con, line_id, lane, claimed_amount, scenario_pair="Budget->Actual"):
    """Verifier(§6.1): 임의 주장이 데이터로 뒷받침되는지. 환각 차단 시연용."""
    row = con.execute("SELECT amount FROM variance_lane WHERE line_id=? AND lane=? AND scenario_pair=?",
                      (line_id, lane, scenario_pair)).fetchone()
    if not row: return {"grounded": False, "reason": "해당 레인 데이터 없음"}
    ok = abs(row[0] - claimed_amount) < 1
    return {"grounded": ok, "actual": row[0], "claimed": claimed_amount,
            "reason": "일치" if ok else "수치 불일치 → 차단(환각)"}


# ============================================================================
# SECTION 25 — 에이전트 관측 / SLO (§6.7: 큐·게이트·충족·비용)
# ============================================================================
def compute_slos(con, asof="2026-06-14"):
    """파이프라인 건강도 지표 + SLO 상태(데이터 품질이 아니라 *에이전트 운영*)."""
    def one(q, *p): return con.execute(q, p).fetchone()[0]
    qdepth = one("SELECT COUNT(*) FROM work_item WHERE status IN('pending','awaiting_playbook')")
    deadletter = one("SELECT COUNT(*) FROM work_item WHERE status='dead'")
    applied = one("SELECT COUNT(*) FROM decision_analysis WHERE status='applied'")
    review = one("SELECT COUNT(*) FROM decision_analysis WHERE status IN('review','provisional')")
    stale = one("SELECT COUNT(*) FROM decision_analysis WHERE status='stale'")
    auto_rate = applied / max(1, applied + review)
    req_total = one("SELECT COUNT(*) FROM request_register")
    req_fulfilled = one("SELECT COUNT(*) FROM request_register WHERE status='fulfilled'")
    fulfil_rate = req_fulfilled / max(1, req_total)
    pending_appr = one("SELECT COUNT(*) FROM task_card WHERE status='pending_approval'")
    runs = one("SELECT COUNT(*) FROM ops_run")
    run_ok = one("SELECT COUNT(*) FROM ops_run WHERE status='completed'")
    worker_succ = run_ok / max(1, runs)
    cost = one("SELECT COALESCE(SUM(cost),0) FROM ops_run")
    # (metric, value, target, status)
    rows = [
        ("queue_depth", qdepth, "≤20", "ok" if qdepth <= 20 else "BREACH"),
        ("dead_letter", deadletter, "0", "ok" if deadletter == 0 else "BREACH"),
        ("auto_apply_rate", f"{auto_rate:.0%}", "monitor", "info"),
        ("review_backlog", review, "≤25", "ok" if review <= 25 else "BREACH"),
        ("stale_decisions", stale, "monitor", "info"),
        ("request_fulfilment", f"{fulfil_rate:.0%}", "≥80%", "ok" if fulfil_rate >= 0.8 else "watch"),
        ("pending_approvals", pending_appr, "clear daily", "ok" if pending_appr <= 20 else "watch"),
        ("worker_success", f"{worker_succ:.0%}", "≥99%", "ok" if worker_succ >= 0.99 else "watch"),
        ("run_cost", f"${cost:.2f}", "budget", "info"),
    ]
    for rs, sla in (("REB_office_rent", 120), ("ECOS_ibr", 40)):   # 참조데이터 freshness 게이팅
        fr = ref_freshness(con, rs, sla, asof)
        rows.append((f"ref_freshness:{rs}", fr.get("status", "-"), f"≤{sla}d",
                     "ok" if fr.get("status") == "fresh" else ("info" if fr.get("status") == "missing" else "watch")))
    return rows


# ============================================================================
# SECTION 26 — 노트 분류·라우팅 (라이프사이클 §2: 단일 인박스 → 3갈래)
# ============================================================================
# 결정론 규칙(키워드/frontmatter). LLM 보강은 동일 시그니처의 구현 교체 지점(인터페이스 불변).
DOMAIN_KEYWORDS = {
    "고정비-차량": ["트럭", "차량", "fleet", "배송차", "지게차", "리스차"],
    "고정비-부동산": ["임차", "임대", "건물", "lease", "물류센터", "사무실", "보증금", "갱신"],
    "고정비-유틸리티": ["전기", "수도", "유틸리티", "가스", "통신요금"],
    "고정비-설비": ["설비", "기계", "분류기", "sorter", "컨베이어", "감가", "손상"],
}
ACTION_VERBS = ["요청", "회신", "확인", "검토", "발송", "예정", "협상", "처리", "보고"]
SENSITIVE_HINTS = ["₩", "만원", "억원", "보증금", "계약금액", "임대료", "단가", "연봉"]
REFERENCE_HINTS = ["방법론", "정의", "ifrs", "ias", "rics", "기준서", "표준", "원칙"]
MEETING_HINTS = ["회의", "미팅", "참석", "안건", "agenda", "논의", "킥오프"]

def classify_note(fm, body):
    """frontmatter+본문 신호로 type/topic/sensitivity/route/tags 분류."""
    text = (body or "").lower(); raw = body or ""; tags = []
    topic = fm.get("topic")
    topic = (topic[0] if isinstance(topic, list) else topic) if topic else None
    if not topic:
        for dom, kws in DOMAIN_KEYWORDS.items():
            if any(k.lower() in text for k in kws): topic = dom; break
    if topic: tags.append(topic)
    sens = fm.get("sensitivity") or ("S1" if any(h.lower() in text for h in SENSITIVE_HINTS) else "S0")
    if sens == "S1": tags.append("기밀")
    t = fm.get("type")
    if not t:
        if any(h in text for h in MEETING_HINTS): t = "meeting"
        elif any(h in text for h in REFERENCE_HINTS): t = "reference"
        elif any(v in raw for v in ACTION_VERBS): t = "action"
        else: t = "personal"
    fcst_line = fm.get("fcst_line")
    if t in ("action", "data") and (fcst_line or (topic or "").startswith("고정비")):
        route = "pipeline"
    elif t == "meeting": route = "meeting_extract"
    elif t == "reference": route = "DA_index"
    else: route = "vault_only"
    conf = 0.8 if (fm.get("type") or topic) else 0.55
    return {"type": t, "topic": topic, "fcst_line": fcst_line, "sensitivity": sens,
            "route": route, "confidence": round(conf, 2), "tags": tags}

def route_note(con, note_id, fm, body, now="2026-06-14", confirmed=False):
    """분류 → note_register 적재 → 라우팅. pipeline행은 사람 확인 후에만 work_item."""
    c = classify_note(fm, body)
    con.execute("INSERT OR REPLACE INTO note_register VALUES(?,?,?,?,?,?,?,?,?,?)",
                (note_id, c["type"], c["topic"], c["fcst_line"], c["sensitivity"], c["route"],
                 c["confidence"], json.dumps(c["tags"], ensure_ascii=False), "routed",
                 fm.get("captured_at", now)))
    out = {"note_id": note_id, **c}
    if c["route"] == "pipeline":
        if confirmed:   # 사람 확인 = 의도적 행위(request 발행에 준함) → correlated work_item
            tok = open_request(con, f"REQ-NOTE-{note_id}", c["fcst_line"], "fixedcost-fpna", now, now=now)
            out["pipeline"] = enqueue_work_item(
                con, "note", f"obsidian://{note_id}", f"note:{note_id}",
                correlation_token=tok,
                payload={"kind": "note", "fcst_line": c["fcst_line"], "topic": c["topic"], "body": body}, now=now)
        else:
            out["pipeline"] = "검토 대기(사람 확인 후 work_item)"
    elif c["route"] == "DA_index":
        out["da"] = "DA 지식베이스 인덱싱(30_Resources)"
    elif c["route"] == "meeting_extract":
        out["meeting"] = "50_Meetings 구조화 추출(결정/액션/예산영향)"
    con.commit()
    return out


# ============================================================================
# SECTION 27 — 산출물 폴더링·매니페스트 (라이프사이클 §5: 출력 택소노미·버전·발행)
# ============================================================================
from datetime import date as _date

def output_path(period, audience, atype, scope, version, run_id, ext, draft=False):
    """출력 택소노미: /reports/<YYYY>/<YYYY-MM>/<audience|_drafts>/<period_scope_type_vN__run>.ext"""
    folder = "_drafts" if draft else audience
    name = f"{period}_{scope}_{atype}_v{version}__{run_id}.{ext}"
    return f"/fpna-fixedcost/reports/{period[:4]}/{period}/{folder}/{name}"

def build_manifest(con, run_id, inputs, now="2026-06-14"):
    """재현 매니페스트: 참조 버전 + 모델 버전 + ledger 규모(어느 시점·어느 지표로 만들었나)."""
    ref_versions = {r[0]: r[1] for r in con.execute("SELECT ref_set,version FROM ref_snapshot")}
    models = sorted({r[0] for r in con.execute(
        "SELECT DISTINCT model_version FROM decision_analysis WHERE model_version IS NOT NULL")})
    return {"run_id": run_id, "built_at": now, "inputs": inputs, "ref_versions": ref_versions,
            "model_versions": models,
            "ledger_decisions": con.execute("SELECT COUNT(*) FROM decision_analysis").fetchone()[0]}

def register_artifact(con, period, audience, atype, scope, run_id, ext, inputs, now="2026-06-14"):
    """산출물을 초안(_drafts)으로 등록 + 매니페스트. 버전은 동일 키 최대+1."""
    ver = con.execute("""SELECT COALESCE(MAX(version),0)+1 FROM artifact_register
        WHERE period=? AND audience=? AND atype=? AND scope=?""", (period, audience, atype, scope)).fetchone()[0]
    aid = f"{period}_{scope}_{atype}_v{ver}__{run_id}"
    path = output_path(period, audience, atype, scope, ver, run_id, ext, draft=True)
    manifest = build_manifest(con, run_id, inputs, now)
    con.execute("INSERT OR REPLACE INTO artifact_register VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (aid, period, audience, atype, scope, ver, run_id, path, ext, "draft",
                 json.dumps(manifest, ensure_ascii=False), now))
    con.commit()
    return {"artifact_id": aid, "status": "draft", "version": ver, "path": path}

def publish_artifact(con, artifact_id, approver, now="2026-06-14"):
    """승인 발행: _drafts → audience 폴더, 발행본 불변(재현 가능). 재발행 시 새 버전."""
    row = con.execute("""SELECT period,audience,atype,scope,version,run_id,ext,status
        FROM artifact_register WHERE artifact_id=?""", (artifact_id,)).fetchone()
    if not row: raise ValueError("artifact 없음")
    period, audience, atype, scope, ver, run_id, ext, status = row
    if status == "published":
        return {"artifact_id": artifact_id, "status": "already_published(immutable)", "action": "재발행은 새 run_id"}
    pub = output_path(period, audience, atype, scope, ver, run_id, ext, draft=False)
    con.execute("UPDATE artifact_register SET status='published', path=? WHERE artifact_id=?", (pub, artifact_id))
    con.commit()
    return {"artifact_id": artifact_id, "status": "published", "path": pub, "by": approver, "immutable": True}


# ============================================================================
# SECTION 28 — 외부 검토 반영 개선 (Splink·promptfoo/DeepEval·FP&A-agent 벤치마크)
# ============================================================================
# 검토 결론: 아키텍처 방향은 분야 베스트프랙티스(HITL·grounded·중요성 변화검토·관측)와 일치.
# 개선 3건: (A) ER u 직접추정(Splink식), (B) eval baseline·비용 회귀, (C) 예측 정확도 추적.
from collections import Counter

# --- 28A. ER: Splink식 u 직접추정 (하드코딩 프라이어 → 데이터 보정 경로) ---
def calibrate_u_from_data(records, fields=("biz_no", "name")):
    """비매칭 우세(스케일) 가정하 값 빈도로 u(비매칭 시 우연 일치 확률) 추정.
    Splink 권고=u 직접추정·m은 EM. 소표본은 프라이어 권장(조건부 독립 가정 유의)."""
    u = {}
    for f in fields:
        vals = [str(r.get(f)) for r in records if r.get(f)]
        if not vals: u[f] = None; continue
        cnt = Counter(vals); tot = len(vals)
        u[f] = round(min(0.99, max(1e-6, sum((c / tot) ** 2 for c in cnt.values()))), 6)
    return {"u_probabilities": u, "n": len(records),
            "note": "스케일(비매칭 우세)에서 신뢰; 소표본은 프라이어; 조건부독립 가정 유의."}

def fs_match_weight(m, u, agree):
    """Fellegi-Sunter 부분 가중치 = log2(m/u)[일치] | log2((1-m)/(1-u))[불일치]. λ 프라이어는 별도."""
    u = min(max(u, 1e-6), 1 - 1e-6); m = min(max(m, 1e-6), 1 - 1e-6)
    return math.log2(m / u) if agree else math.log2((1 - m) / (1 - u))

def fs_prior_weight(lam):
    """λ(임의 두 레코드가 매치일 사전확률) → 사전 매치가중치 M_prior."""
    lam = min(max(lam, 1e-9), 1 - 1e-9)
    return math.log2(lam / (1 - lam))

# --- 28B. eval: 정적 임계 + 직전 baseline 대비 회귀 + 비용 회귀 (promptfoo/DeepEval 패턴) ---
def eval_regression_vs_baseline(con, suite, value, le=False, margin=0.02):
    """직전 동일 suite 결과 대비 악화 감지(정적 임계가 못 잡는 점진 퇴행)."""
    prev = con.execute("SELECT value FROM eval_run WHERE suite=? ORDER BY id DESC LIMIT 1", (suite,)).fetchone()
    if not prev: return {"suite": suite, "baseline": None, "regressed": False}
    base = prev[0]
    regressed = (value > base + margin) if le else (value < base - margin)
    return {"suite": suite, "baseline": base, "current": value, "regressed": regressed}

def eval_cost_regression(con, budget):
    """비용 회귀(Gartner 실패모드 #1=비용 폭주). ops_run 누적 비용 vs 예산."""
    cost = con.execute("SELECT COALESCE(SUM(cost),0) FROM ops_run").fetchone()[0]
    return {"metric": "run_cost", "value": round(cost, 4), "budget": budget, "regressed": cost > budget}

# --- 28C. 예측 정확도/편향 추적 (분야가 지목하는 핵심 열화 신호) ---
def record_forecast_actual(con, line, period, forecast, actual, now="2026-06-14"):
    con.execute("INSERT INTO forecast_actual VALUES(?,?,?,?,?)", (line, period, forecast, actual, now)); con.commit()

def forecast_accuracy(con, line=None, mape_threshold=0.10):
    """과거 fcst vs 실현 actuals로 MAPE·bias. 임계 초과 시 degraded(시간 경과 열화 감시)."""
    q = "SELECT forecast,actual FROM forecast_actual" + (" WHERE line=?" if line else "")
    rows = con.execute(q, (line,) if line else ()).fetchall()
    pts = [(f, a) for f, a in rows if a]
    if not pts: return {"line": line or "ALL", "n": 0, "status": "no_data"}
    mape = sum(abs(f - a) / abs(a) for f, a in pts) / len(pts)
    bias = sum((f - a) / abs(a) for f, a in pts) / len(pts)
    return {"line": line or "ALL", "n": len(pts), "mape": round(mape, 4), "bias": round(bias, 4),
            "status": "degraded" if mape > mape_threshold else "ok"}


# ============================================================================
# SECTION 29 — 구현 보강·정밀화 (Splink EM/TF · CoVe 하이브리드 검증 · 외부 어댑터 · ABC · 발송 sink · 설정)
# ============================================================================
import os, re

# --- B6: ER m을 EM으로 학습 (Splink: u 직접추정 + m EM). 비라벨 페어 베르누이 혼합 ---
def train_er_em(records, field="biz_no", iters=20, init_m=0.95, init_u=0.001, init_lambda=0.001):
    pairs = []
    for i in range(len(records)):
        for j in range(i + 1, len(records)):
            a, b = records[i].get(field), records[j].get(field)
            pairs.append(1 if (a and b and a == b) else 0)
    if not pairs: return {"m": init_m, "u": init_u, "lambda": init_lambda, "n_pairs": 0}
    m, u, lam = init_m, init_u, init_lambda
    for _ in range(iters):
        ws = []
        for g in pairs:
            pm = (m if g else 1 - m) * lam
            pu = (u if g else 1 - u) * (1 - lam)
            ws.append(pm / (pm + pu) if (pm + pu) > 0 else 0.0)
        sw = sum(ws); npr = len(pairs)
        lam = sw / npr
        m = (sum(w * g for w, g in zip(ws, pairs)) / sw) if sw > 0 else m
        u = (sum((1 - w) * g for w, g in zip(ws, pairs)) / (npr - sw)) if (npr - sw) > 0 else u
        m, u, lam = min(max(m, 1e-4), 1 - 1e-4), min(max(u, 1e-6), 1 - 1e-6), min(max(lam, 1e-6), 1 - 1e-6)
    return {"m": round(m, 4), "u": round(u, 6), "lambda": round(lam, 4), "n_pairs": len(pairs)}

# --- A2: 하이브리드 클레임 검증 (CoVe/attribution식: 수치 일치 + 어휘 정렬) ---
def verify_claim(evidence_text, claim_text, lexical_min=0.6):
    """수치 있으면 수치가 근거에 존재해야 함; 없으면 핵심 토큰 포함율로 판정(환각 차단)."""
    nums_c = re.findall(r"-?\d+(?:\.\d+)?", claim_text.replace(",", ""))
    if nums_c:
        nums_e = set(re.findall(r"-?\d+(?:\.\d+)?", evidence_text.replace(",", "")))
        mags_e = {x.lstrip("-") for x in nums_e}
        ok = all(n in nums_e or n.lstrip("-") in mags_e for n in nums_c)  # 부호 차이 허용(방향은 어휘로)
        return {"grounded": ok, "mode": "numeric", "reason": "수치 일치" if ok else "수치 불일치 → 차단(환각)"}
    toks = [t for t in re.findall(r"\w+", claim_text.lower()) if len(t) > 1]
    ev = evidence_text.lower()
    cover = sum(1 for t in toks if t in ev or any(t[:k] in ev for k in (len(t)-1,) if k >= 2)) / max(1, len(toks))
    ok = cover >= lexical_min
    return {"grounded": ok, "mode": "lexical", "coverage": round(cover, 2),
            "reason": f"어휘 포함 {cover:.0%}" + ("" if ok else " → 차단")}

# --- B3/A4: 외부 데이터 어댑터 (fetcher 인터페이스; 폐쇄망=file, 인터넷=api 동일 시그니처) ---
def file_fetcher(path):
    """폐쇄망 기본: 승인 구간이 SharePoint에 착지시킨 스냅샷 파일을 읽음(오프라인 동작)."""
    with open(path, encoding="utf-8") as f: return json.load(f)

def api_fetcher(url, env_key="GODATA_API_KEY"):
    """인터넷 구간: 환경변수 키로 OpenAPI 호출(REB/ECOS/KOSIS). file_fetcher와 동일 반환."""
    key = os.environ.get(env_key)
    if not key: raise RuntimeError(f"{env_key} 미설정 — 폐쇄망에서는 file_fetcher 사용")
    import urllib.request
    with urllib.request.urlopen(f"{url}?serviceKey={key}", timeout=20) as r:  # 승인 프록시 경유
        return json.loads(r.read().decode())

def fetch_and_ingest(con, table, keys, ref_set, source, url, license, version, fetcher, now="2026-06-14"):
    """fetcher가 가져온 행을 reference_data(SCD2)로 적재. 엔진은 이후 get_ibr/regional_params로 조회."""
    rows = fetcher()
    return ingest_snapshot(con, table, rows, keys, ref_set, source, url, license, version, now)

# --- B7: 활동기준원가(ABC) 일반 배부 (driver 비중으로 cost pool 배분) ---
def allocate_cost_abc(pool_amount, drivers):
    """drivers={cost_object: driver_qty}. 합계 가중으로 배분(잔차는 최대 객체에 보정→합 보존)."""
    total = sum(drivers.values())
    if total <= 0: return {k: 0.0 for k in drivers}
    alloc = {k: round(pool_amount * q / total, 2) for k, q in drivers.items()}
    drift = round(pool_amount - sum(alloc.values()), 2)
    if drift and alloc:
        big = max(alloc, key=lambda k: alloc[k]); alloc[big] = round(alloc[big] + drift, 2)
    return alloc

# --- B2: 발송 sink (no-op 아님). 기본=sent_log 기록; 외부(COM/GL)는 동일 시그니처로 교체 ---
def log_sender(con, task_id, idem, side_effect, now="2026-06-14"):
    con.execute("INSERT OR IGNORE INTO sent_log(task_id,idempotency_key,channel,detail,sent_at) VALUES(?,?,?,?,?)",
                (task_id, idem, "log", side_effect, now))
    return f"log::{idem}"
SENDERS = {"log": log_sender}  # sender=통지 채널(메일/Teams). 'outlook_com'을 같은 시그니처로 등록해 교체. GL 기록 경로 없음(단방향)

# --- C1: 완전 IBR 매트릭스 기본값 (KRW; ECOS 회사채AA-/국고채 + 담보 스프레드). 플레이스홀더 아님 ---
def seed_full_ibr_matrix(con, version="2026-06", now="2026-06-14"):
    rows = [
        {"currency": "KRW", "term_band": "1-3y", "security": "secured",   "ibr": 0.047, "benchmark": "국고채3y+α", "spread": 0.012},
        {"currency": "KRW", "term_band": "1-3y", "security": "unsecured", "ibr": 0.055, "benchmark": "회사채AA-3y", "spread": 0.010},
        {"currency": "KRW", "term_band": "3-5y", "security": "secured",   "ibr": 0.052, "benchmark": "회사채AA-3y", "spread": 0.004},
        {"currency": "KRW", "term_band": "3-5y", "security": "unsecured", "ibr": 0.061, "benchmark": "회사채AA-3y", "spread": 0.013},
        {"currency": "KRW", "term_band": "5y+",  "security": "secured",   "ibr": 0.058, "benchmark": "회사채AA-5y", "spread": 0.006},
        {"currency": "KRW", "term_band": "5y+",  "security": "unsecured", "ibr": 0.067, "benchmark": "회사채AA-5y", "spread": 0.015},
    ]
    return ingest_snapshot(con, "ibr_matrix", rows, ["currency", "term_band", "security"],
                           "ECOS_ibr", "한국은행 ECOS", "ecos.bok.or.kr", "KOGL-1", version, now)

# --- C2~C8: 운영 기본값 (작동하는 기본; 오버라이드 선택) ---
DEFAULT_HURDLE = 0.09                                    # C2 자본 의사결정 할인율
DEFAULT_CGU = {"name": "DC_sorting_cgu", "test_trigger": "indicator|annual",
               "impairment_threshold": 0.0}             # C3 (회수가능<장부 시 손상)
HOUSE_STYLE = {                                          # C6 보고 톤/양식 통일
    "title_rule": "action_title",                        # 제목=결론 문장
    "structure": ["message_title", "key_exhibit", "driver_bridge", "source_footnote"],
    "tone": "concise_executive", "numbers_from": "L2_ledger", "grounding": "required"}
RETENTION_POLICY = {"boardpack": "7y", "variance": "7y", "fcst": "3y",      # C7
                    "draft": "90d", "note_vault": "indefinite", "audit_evidence": "7y"}
PII_POLICY = {"S1": {"sync": "off_device", "access": "need_to_know"},        # C8
              "default_mask": True, "retention": "per_class"}
TAXONOMY_DEFAULT = {"audiences": ["CEO", "controller", "internal"],          # C7
                    "types": ["boardpack", "variance", "fcst", "memo"]}


# ============================================================================
# SECTION 30 — 산출물 콘텐츠 스펙 (B4: grounded·house_style; 렌더는 BIGS/academic-slide)
# ============================================================================
# PPT 엔진(pptxgenjs/python-pptx)은 stdlib+openpyxl 제약 밖이고, 사내 BIGS deck_system·
# academic-slide 스킬이 렌더 툴체인이다. 이 시스템의 책임=콘텐츠(결론 제목·exhibit·근거 각주·
# 그라운딩)를 구조화해 그 렌더러에 넘기는 것. 콘텐츠는 ledger/variance/내러티브에서 생성.
def _evidence_health(con):
    rows = con.execute("SELECT grounded,approval_status FROM control_evidence").fetchall()
    n = len(rows); grounded = sum(1 for g, _ in rows if g); appr = sum(1 for _, a in rows if a == "approved")
    return {"controls": n, "grounded": grounded, "approved": appr}

def build_board_deck_spec(con, line, period="2026-06", audience="CEO"):
    """CEO 보고 덱 콘텐츠 스펙. 각 슬라이드=결론 제목(house_style)+exhibit+근거 각주+그라운딩.
    결론 제목(텍스트 주장)은 verify_claim(수치+어휘 하이브리드)으로 근거 검증 후 게이팅.
    렌더는 BIGS deck_system/academic-slide 스킬이 수행(이 함수는 콘텐츠만)."""
    nv = narrate_variance(con, line)
    lanes = [(ln, amt, ev) for ln, amt, ev in con.execute(
        "SELECT lane,amount,evidence FROM variance_lane WHERE line_id=?", (line,))]
    total = sum(a for _, a, _ in lanes)
    # 근거 텍스트: 레인명·금액·결정 라벨 → 결론 제목을 이 근거로 검증
    evidence = f"{line} 합계 {round(total)} " + " ".join(
        f"{ln} {round(a)} {ev}" for ln, a, ev in lanes)
    eh = _evidence_health(con)
    direction = "감소" if total < 0 else "증가"
    top = sorted([(ln, a) for ln, a, _ in lanes], key=lambda x: -abs(x[1]))
    t_title = f"고정비 예산 대비 ₩{abs(total):,.0f} {direction}, 주요인 {top[0][0] if top else '-'}"
    slides = [
        {"layout": "title", "action_title": f"{period} 고정비 리뷰 — {line}",
         "subtitle": "계약 기반 결정·투영·근거", "source": "L2 ledger"},
        {"layout": "stat_bridge", "action_title": t_title,
         "exhibit": {"type": "waterfall", "data": [(ln, a) for ln, a, _ in lanes]},
         "narrative": nv["narrative"], "source": f"variance_bridge::{line}",
         "grounded": verify_claim(evidence, t_title)["grounded"]},     # 하이브리드 검증 게이트
        {"layout": "evidence_status",
         "action_title": f"근거·통제: {eh['grounded']}/{eh['controls']} grounded, {eh['approved']} 승인",
         "exhibit": {"type": "icfr_table", "data": eh}, "source": "control_evidence"},
    ]
    return {"audience": audience, "house_style": HOUSE_STYLE,
            "grounded": nv["all_grounded"] and all(s.get("grounded", True) for s in slides),
            "renderer": "BIGS deck_system | academic-slide", "slides": slides}

def build_report_spec(con, line, period="2026-06", fmt="pdf"):
    """정식 보고서(PDF/문서) 콘텐츠 스펙. 동일 그라운딩·house_style; 렌더는 해당 툴체인."""
    deck = build_board_deck_spec(con, line, period, audience="controller")
    return {"format": fmt, "house_style": HOUSE_STYLE, "grounded": deck["grounded"],
            "sections": [{"heading": s["action_title"], "body": s.get("narrative", ""),
                          "exhibit": s.get("exhibit"), "source": s.get("source")} for s in deck["slides"]],
            "retention": RETENTION_POLICY.get("boardpack")}


# ============================================================================
# SECTION 19 — 데모 / 실행 검증
# ============================================================================
def _line(t): print("\n" + "=" * 74 + f"\n{t}\n" + "=" * 74)

def main():
    con = init_db()

    # ----- 0) Reference Data 적재 (외부 어댑터가 SharePoint에 착지한 스냅샷 → reference_data) -----
    _line("0) Reference Data (REB 지역 임대지표 + ECOS IBR) 적재 → 엔진이 조회")
    print("  REB 스냅샷:", seed_reb_snapshot(con, version="2026Q1"))
    print("  ECOS IBR  :", seed_ecos_ibr(con, version="2026-06"))
    rp = regional_params(con, "판교", 4, "2026Q1")
    ibr = get_ibr(con, "KRW", "3-5y", "secured")
    print(f"  regional_params(판교,grade4): anchor ₩{rp['market_rent_anchor']:,.0f}/㎡, 월추세 {rp['market_trend_pm']:.4%} (REB 지수 유도)")
    print(f"  get_ibr(KRW,3-5y,secured): {ibr['ibr']:.3%} (벤치마크 {ibr['benchmark']})")
    print("  freshness:", ref_freshness(con, "REB_office_rent", sla_days=120, asof="2026-06-14"),
          ref_freshness(con, "ECOS_ibr", sla_days=40, asof="2026-06-14"))

    # ----- 1) 건물 임차 유불리 → 갱신 가정 투영 -----
    _line("1) 건물 임차 유불리 + 시장임대료 추정 (RICS 비교법) → 갱신 투영")
    subject = Subject("이천", 12000, 60, 3, "net", 9800)
    P = lambda i, a: Provenance(f"sp://comps/2026/{i}", a, "2026-05-01")
    comps = [Comp("C1", 9000, 3, 60, 4, 11000, 3, "net", True, "letting", P(1, Authority.INVOICE_PO)),
             Comp("C2", 8600, 6, 60, 10, 13000, 3, "net", True, "letting", P(2, Authority.INVOICE_PO)),
             Comp("C3", 10200, 0, 36, 2, 8000, 4, "gross", False, "quoting", P(3, Authority.BROKER_QUOTE)),
             Comp("C4", 8800, 4, 60, 18, 12500, 3, "net", True, "letting", P(4, Authority.INVOICE_PO))]
    rp_ic = regional_params(con, "이천", 3, "2026Q1")
    params = CompAdjustParams(market_trend_pm=(rp_ic["market_trend_pm"] if rp_ic else 0.0))
    res, inp = estimate_market_rent(subject, comps, params=params, abs_materiality=50_000_000)
    renewal_annual = res.recommended_value * subject.size_sqm * 12   # 시장 목표가로 갱신 가정
    proj = project_fcst_lines("lease_favorability", renewal_payment=renewal_annual, renewal_term=5)
    print(f"시장임대료 ₩{res.recommended_value:,.0f}/㎡ (80%CI {res.conf_interval[0]:,.0f}~{res.conf_interval[1]:,.0f}) | 권고: {res.recommendation}")
    print("게이트:", submit(con, "이천 임차 갱신 유불리", res, inp, projections=proj))

    # ----- 2) 트럭 구매 vs 리스 → 선택안 투영 -----
    _line("2) 트럭 구매 vs 리스 (세후 DCF + IFRS16) → 선택안 fcst 투영")
    own = OwnPlan(120_000_000, 8, 20_000_000, 14_000_000, 0.22, 1_000_000)
    ibr_v = get_ibr(con, "KRW", "3-5y", "secured")["ibr"]
    lease = LeasePlan(21_000_000, 6, ibr_v, escalation=0.02, annual_service_in_payment=2_000_000)
    pv = Provenance("sp://capex/quote.pdf", Authority.BROKER_QUOTE, "2026-05-20")
    res2, inp2 = analyze_buy_vs_lease(own, lease, hurdle_rate=0.09, horizon=8, abs_materiality=10_000_000, prov_own=pv, prov_lease=pv)
    proj2 = project_fcst_lines("buy_vs_lease", own=own, lease=lease, recommendation=res2.recommendation)
    print(f"권고: {res2.recommendation}")
    print(f"  {res2.notes[0]}")
    print(f"  {res2.notes[2]}")   # NAL(등가대출) — 재무이론 관점 병기
    print("게이트:", submit(con, "이천 DC 트럭 조달", res2, inp2, projections=proj2))
    if proj2:
        lt, sched, note = proj2[0]
        print(f"  투영[{lt}] ({note}) 1~3기:")
        for r in sched[:3]: print("   ", r)

    # ----- 3) 기계 손상 → 손상후 수정 상각 투영 -----
    _line("3) 자동분류기 손상 (IAS 36) → 손상후 수정 상각 투영 + SOX 증적")
    cgu = CGU("DC이천-분류라인",
              [CGUAsset("sorter-01", 5_000_000_000, life_years=10, elapsed_years=3, residual=200_000_000),
               CGUAsset("conveyor-01", 1_200_000_000, life_years=8, elapsed_years=3, residual=50_000_000)],
              [900_000_000, 850_000_000, 800_000_000, 750_000_000, 700_000_000],
              0.13, terminal_growth=0.01, fair_value=4_300_000_000, costs_of_disposal=150_000_000)
    pvf = Provenance("sp://appraisal/sorter.pdf", Authority.SIGNED_CONTRACT, "2026-06-01")
    res3, inp3 = test_impairment(cgu, abs_materiality=200_000_000, prov_fv=pvf)
    proj3 = project_fcst_lines("impairment", cgu=cgu, alloc=res3._alloc)
    print(f"VIU=₩{value_in_use(cgu):,.0f} | FVLCD=₩{fvlcd(cgu):,.0f} | 장부=₩{cgu.carrying_total:,.0f}")
    print(f"권고: {res3.recommendation}")
    print(f"headroom 민감도: {res3.sensitivity}")
    print("게이트:", submit(con, "분류라인 분기 손상검토", res3, inp3, projections=proj3))
    for lt, sched, note in proj3:
        print(f"  투영[{lt}] ({note}) 1~3기:")
        for r in sched[:3]: print("   ", r)

    # ----- 4) 트리거/큐: 자료요청 발송 → 회신 인입 → 라우팅 → 워커 → 결정 -----
    _line("4) 트리거/큐 (§4/§6.2/§6.3): 요청 발송 → 회신 인입 → 드레인 → 워커")
    token = open_request(con, "REQ-001", "HQ_lease_pangyo", "facilities", "2026-06-20")
    print("  요청 발송:", token)
    reply = {"kind": "lease_comps",
             "subject": {"location": "판교", "size_sqm": 8000, "term_months": 60, "grade": 4, "cost_basis": "net", "contract_rent_per_sqm": 21000},
             "comps": [
                 {"comp_id": "K1", "headline_rent_per_sqm": 19500, "rent_free_months": 3, "term_months": 60, "months_ago": 3, "size_sqm": 7500, "grade": 4, "cost_basis": "net", "same_location": True, "txn_type": "letting", "uri": "sp://reply/REQ-001/k1", "auth": "INVOICE_PO", "asof": "2026-06-15"},
                 {"comp_id": "K2", "headline_rent_per_sqm": 18800, "rent_free_months": 4, "term_months": 60, "months_ago": 8, "size_sqm": 9000, "grade": 4, "cost_basis": "net", "same_location": True, "txn_type": "letting", "uri": "sp://reply/REQ-001/k2", "auth": "INVOICE_PO", "asof": "2026-06-15"},
                 {"comp_id": "K3", "headline_rent_per_sqm": 20100, "rent_free_months": 2, "term_months": 48, "months_ago": 2, "size_sqm": 6800, "grade": 4, "cost_basis": "net", "same_location": True, "txn_type": "letting", "uri": "sp://reply/REQ-001/k3", "auth": "INVOICE_PO", "asof": "2026-06-15"}],
             "materiality": 50_000_000}
    e1 = enqueue_work_item(con, "outlook", "msg://reply1", "imsgid-AAA", correlation_token=token, payload=reply)
    e2 = enqueue_work_item(con, "outlook", "msg://reply1", "imsgid-AAA", correlation_token=token, payload=reply)  # 중복
    print("  인입:", e1, "| 동일 재인입(dedup):", e2["enqueued"])
    enqueue_work_item(con, "outlook", "msg://spam", "imsgid-ZZZ", correlation_token=None, payload={"kind": "unknown"})  # unsolicited
    for d in drain(con):
        print("  드레인:", {k: d[k] for k in d if k != "result"}, "| 결정:", d.get("result", {}).get("recommendation") if isinstance(d.get("result"), dict) else d.get("result"))

    # ----- fcst 라인 적재 확인 -----
    _line("적재된 fcst 라인 투영 (SQLite, 발췌)")
    for row in con.execute("SELECT analysis_id,line_type,period,value FROM fcst_line_projection ORDER BY analysis_id,period LIMIT 10"):
        print("  fcst:", row)

    # ----- ICFR (승인 전) -----
    _line("ICFR/SOX 통제 요약 — 승인 전")
    for cid, risk, freq, cnt, last, status in icfr_summary(con):
        print(f"  {cid:<10}{status:<18}(evid {cnt})")

    # ----- 승인(SoD) → outbox 적재 -----
    _line("승인 워크플로 (§6.8 SoD) → transactional outbox 적재")
    for (tid,) in con.execute("SELECT id FROM task_card WHERE status='pending_approval' ORDER BY id"):
        print("  승인:", approve_task(con, tid, approver="controller"))
    con.execute("INSERT INTO task_card(analysis_id,task_type,payload,risk_tier,status,requester,approver,created_at) VALUES(1,'x','{}','low','pending_approval','fpna-analyst',NULL,'2026-06-14')")
    bad = con.execute("SELECT MAX(id) FROM task_card").fetchone()[0]
    try: approve_task(con, bad, approver="fpna-analyst")
    except SoDViolation as e: print("  SoD 차단 OK:", e)

    # ----- outbox 실행 (exactly-once: 2회 호출해도 1회만) -----
    _line("process_outbox (§6.3 exactly-once) — 2회 호출(멱등 검증)")
    print("  1차 실행:", len(process_outbox(con)), "건 confirmed")
    print("  2차 실행(멱등):", len(process_outbox(con)), "건 (중복 0)")

    # ----- ICFR (실행 후 → effective) -----
    _line("ICFR/SOX 통제 요약 — 실행 후 (effective 완성)")
    for cid, risk, freq, cnt, last, status in icfr_summary(con):
        print(f"  {cid:<10}{status:<18}(evid {cnt})")

    # ----- 5) Variance Bridge (가정 변경 레인 = 임차 결정 귀속) -----
    _line("5) Variance Bridge — HQ_lease (가정변경 레인 = 결정 투영에서 자동 산출, tie-out)")
    br = build_variance_bridge(con, "HQ_lease_pangyo", budget=2_016_000_000, actual=1_950_000_000,
                               lanes={"indexation": +40_000_000, "one_time": +12_000_000})
    ac = br["lanes"].get("assumption_change", 0.0)
    print(f"  bridge:", {k: (f"₩{v:,.0f}" if isinstance(v, (int, float)) else v) for k, v in br["lanes"].items()})
    print(f"  assumption_change=₩{ac:,.0f} (HQ 갱신 결정의 1차년 P&L − 예산에서 귀속), tie-out={br['tie_out']}")
    print(f"  tie-out (budget+Σlane−actual) = {br['tie_out']} → {'PASS(=0)' if br['tie_out'] == 0 else 'FAIL'}")
    ac = con.execute("SELECT evidence FROM variance_lane WHERE line_id=? AND lane='assumption_change'", ("HQ_lease_pangyo",)).fetchone()
    print("  assumption_change 근거(결정):", ac[0] if ac else "-")

    # ----- 6) 신뢰 감쇠 (§7.7): 노후 시장추정 → stale → 재확인 요청 -----
    _line("6) 시간 기반 신뢰 감쇠 (§7.7) — 9개월 전 시장추정의 confidence 감쇠")
    con.execute("""INSERT INTO decision_analysis(question,domain,fcst_line,evidence_kind,inputs,model,model_version,
            result,confidence,conf_interval,sensitivity,materiality_band,grounded,status,linked_assumption_card,next_review,created_at)
            VALUES('과거 시장임대료 추정','lease_favorability','DC부산_lease','market_estimate','[]','re_comparable','re_comparable/1.0.0',
            '{\"recommendation\":\"유지\"}',0.72,'[0,0]','{}','n/a',1,'applied','assumption::lease_favorability',NULL,'2025-09-01')""")
    print("  감쇠 처리:", apply_confidence_decay(con, asof="2026-06-14", stale_threshold=0.4))
    rc = con.execute("SELECT analysis_id,task_type,status FROM task_card WHERE task_type='propose_data_request'").fetchall()
    print("  자동 생성된 재확인 Task:", rc)

    # ----- 7) 콜드스타트 / 역사 백필 (§7.11): 부임 1일차 부트스트랩 -----
    _line("7) 콜드스타트 부트스트랩 (§7.11) — 기존 계약·과거 GL → provisional 원장")
    contracts = [
        {"name": "본사 임차(판교)", "domain": "lease_favorability", "fcst_line": "HQ_lease_pangyo", "value": 2_016_000_000, "doc": "sp://contracts/hq_lease.pdf", "asof": "2024-03-01"},
        {"name": "자동분류기(이천)", "domain": "impairment", "fcst_line": "DC이천_sorter", "value": 5_000_000_000, "doc": "sp://contracts/sorter_invoice.pdf", "asof": "2023-06-01"},
    ]
    gl_baseline = {"HQ_lease_pangyo": {1: 168_000_000, 2: 168_000_000, 3: 168_000_000}}
    print("  부트스트랩:", bootstrap_from_history(con, contracts, gl_baseline))
    print("  일괄 검토 승인:", bulk_confirm_provisional(con, approver="controller"))

    # ----- 8) 거래처 Entity Resolution (§7.9): 명칭 변형 클러스터링 -----
    _line("8) 거래처 Entity Resolution (§7.9 Fellegi-Sunter) — 명칭 변형 통합")
    vendors = [
        {"id": 1, "name": "쿠팡로지스틱스(주)", "biz_no": "123-45-67890", "addr": "서울 송파구"},
        {"id": 2, "name": "쿠팡로지스틱스 주식회사", "biz_no": "123-45-67890", "addr": "서울 송파"},
        {"id": 3, "name": "Coupang Logistics Co., Ltd.", "biz_no": "123-45-67890", "addr": "Seoul Songpa"},
        {"id": 4, "name": "한진물류(주)", "biz_no": "999-88-77776", "addr": "인천 중구"},
    ]
    er = resolve_entities(con, vendors)
    print("  클러스터:", er["clusters"])
    print("  검토 큐(경계 점수):", er["review_pairs"] or "없음")

    # ----- 9) 계약 개정 감지 (§7.8): 원본 → 부속합의(개정) → 신규 -----
    _line("9) 계약 개정 감지 (§7.8) — 원본 → 부속합의(개정·재스케줄) → 신규")
    orig = {"counterparty_name": "쿠팡로지스틱스(주)", "biz_no": "123-45-67890", "asset_or_property": "HQ_pangyo",
            "contract_no": "L-HQ-001", "monthly_amount": 168_000_000, "term_months": 60, "escalation": 0.02,
            "source_doc": "sp://contracts/L-HQ-001.pdf", "fcst_line": "HQ_lease_pangyo"}
    print("  원본:", detect_contract_change(con, orig))
    amend = {"counterparty_name": "쿠팡로지스틱스 주식회사", "biz_no": "123-45-67890", "asset_or_property": "HQ_pangyo",
             "contract_no": "L-HQ-001", "monthly_amount": 175_000_000, "term_months": 60, "escalation": 0.03,
             "source_doc": "sp://contracts/L-HQ-001-addendum.pdf", "fcst_line": "HQ_lease_pangyo"}
    print("  부속합의(이름 변형, 같은 사업자번호):", detect_contract_change(con, amend))
    new_c = {"counterparty_name": "한진물류(주)", "biz_no": "999-88-77776", "asset_or_property": "DC_busan",
             "contract_no": "L-DC-BUS-007", "monthly_amount": 90_000_000, "term_months": 48, "escalation": 0.02,
             "source_doc": "sp://contracts/L-DC-BUS-007.pdf", "fcst_line": "DC부산_lease"}
    print("  신규 계약:", detect_contract_change(con, new_c))
    print("  contract_master:", [r for r in con.execute("SELECT contract_id,version,amendment_seq,status,monthly_amount FROM contract_master ORDER BY contract_id")])
    # 개정 → rebuild 소비(영향 라인 결정 무효화 + 재예측 Task) — §6.6 seam
    rb = process_rebuild_requests(con)
    print(f"  rebuild 소비: {rb if rb else '대기 없음'}")
    # overdue 트리거: 기한 경과 미충족 요청 → 드레인이 overdue_escalation 처리
    open_request(con, "REQ-OVD", "DC_truck_lease", "fixedcost-fpna", "2026-01-31")  # 과거 기한
    od = drain(con, now="2026-06-14")
    print(f"  overdue 스캔→드레인: {[r for r in od if 'routed' in r and r.get('routed')=='overdue_escalation'] or '처리됨'}")

    # ----- 10) 민감도 / What-if (§7.3): 임차 인상률·갱신여부 / buy-vs-lease 안정성 -----
    _line("10) 민감도/What-if (§7.3) — 원장 불변·발송 없음 (분석 산출물)")
    ls = lease_sensitivity(con, "HQ_lease_pangyo", base_payment=2_100_000_000, base_term_yr=5,
                           base_escalation=0.02, ibr=get_ibr(con)["ibr"],
                           escalations=[0.02, 0.03, 0.05], renew_options=[True, False])
    print(f"  임차 base(2%,갱신) 총 P&L ₩{ls['base_total']:,.0f}")
    for c in ls["cases"]:
        if c["esc"] in (0.05,) and c["renew"]:
            print(f"   다운사이드(5%,갱신): 총 ₩{c['total']:,.0f}, base 대비 +₩{c['delta']:,.0f}")
        if not c["renew"]:
            print(f"   비갱신(이전/철수,esc{c['esc']:.0%}): 총 ₩{c['total']:,.0f}, base 대비 ₩{c['delta']:,.0f}")
    bs = buy_vs_lease_sensitivity(con, "fleet_icheon", own, lease, hurdles=[0.05, 0.07, 0.09, 0.11], horizon=8)
    print(f"  buy-vs-lease base 권고={bs['base_rec']}; 할인율별:")
    for c in bs["cases"]:
        print(f"   hurdle {c['hurdle']:.0%}: {c['rec']}{' (FLIP!)' if c['flips'] else ''} (소유−리스 NPV ₩{c['own_minus_lease']:,.0f})")

    # ----- 11) 지속 eval / 회귀 (§6.9): 배포 게이트 -----
    _line("11) 지속 eval/회귀 (§6.9) — 라우터·게이트보정·그라운딩 → 배포 게이트")
    dg = eval_deploy_gate(con)
    for s in dg["suites"]:
        print(f"  {s['suite']:<18} {s['metric']}={s['value']} ({s['cmp']}{s['threshold']}) → {'PASS' if s['passed'] else 'FAIL'}  {s['detail']}")
    print(f"  배포: {dg['deploy']}" + (f" (실패: {dg['failing']})" if dg["failing"] else ""))

    # ----- 12) Variance 내러티브 (§7.14): 그라운딩 — 환각 차단 -----
    _line("12) Variance 내러티브 생성 (§7.14) — cite-back 그라운딩")
    nv = narrate_variance(con, "HQ_lease_pangyo")
    print("  내러티브:", nv["narrative"])
    print(f"  all_grounded={nv['all_grounded']}, 차단된 문장={nv['blocked'] or '없음'}")
    print("  Verifier(정상 주장):", verify_narrative_claim(con, "HQ_lease_pangyo", "assumption_change", -118_000_000))
    print("  Verifier(환각 주장):", verify_narrative_claim(con, "HQ_lease_pangyo", "assumption_change", -500_000_000))

    # ----- 13) 에이전트 관측 / SLO (§6.7) -----
    _line("13) 에이전트 관측 / SLO (§6.7) — 파이프라인 건강도")
    print(f"  {'metric':<22}{'value':<12}{'target':<14}status")
    for metric, value, target, status in compute_slos(con):
        print(f"  {metric:<22}{str(value):<12}{target:<14}{status}")

    # ----- 14) 노트 분류·라우팅 (라이프사이클 §2): 단일 인박스 → 3갈래 -----
    _line("14) 노트 분류·라우팅 (단일 인박스 → 파이프라인/DA지식/볼트)")
    notes = [
        ("n1", {"captured_at": "2026-06-14T09:00", "source": "voice"},
         "판교 본사 임대인이 신규 임대료 ₩175,000,000/월 조건 회신 예정. 자료요청해서 검토 필요."),
        ("n2", {"source": "text"}, "IFRS16 리스 IBR은 WACC가 아니라 증분차입이자율이라는 점 메모. 방법론 참고."),
        ("n3", {"type": "meeting", "source": "text"}, "월간 고정비 리뷰 회의: 분류기 손상 검토 안건 논의, 액션 3건."),
        ("n4", {"source": "voice"}, "주말에 등산 다녀옴. 날씨 좋았음."),
    ]
    for nid, fm, body in notes:
        conf = (nid == "n1")   # n1은 사람이 확인했다고 가정 → work_item
        r = route_note(con, nid, fm, body, confirmed=conf)
        extra = r.get("pipeline") or r.get("da") or r.get("meeting") or "볼트 보관"
        print(f"  {nid}: type={r['type']:<9} topic={str(r['topic']):<12} sens={r['sensitivity']} → {r['route']:<14} | {extra if not isinstance(extra, dict) else extra}")

    # ----- openpyxl 보고서 (View Contract + Variance Bridge 차트) -----
    _line("openpyxl 보고서 생성 (전체 시간축·recon·evidence health·ICFR·Variance Bridge)")
    try:
        path = build_report(con, "fixed_cost_report.xlsx", periods=8)
        print("  생성:", path)
    except ImportError:
        print("  (openpyxl 미설치 — 보고서 생략. 엔진/카드는 정상)")

    # ----- 15) 산출물 폴더링·발행 (라이프사이클 §5): 초안 → 승인 발행(불변) -----
    _line("15) 산출물 폴더링·매니페스트 (출력 택소노미·버전·발행)")
    a1 = register_artifact(con, "2026-06", "controller", "variance", "HQlease", "r0617", "xlsx",
                           inputs=["L2_ledger", "variance_lane", "REB_2026Q1"])
    print("  등록(초안):", a1["path"])
    pub = publish_artifact(con, a1["artifact_id"], approver="controller")
    print("  발행:", pub["path"], "| immutable:", pub["immutable"])
    print("  재발행 시도(불변):", publish_artifact(con, a1["artifact_id"], approver="controller")["status"])
    a2 = register_artifact(con, "2026-06", "CEO", "boardpack", "fixedcost", "r0617", "pptx",
                           inputs=["variance_bridge", "grounded_narrative", "evidence_health"])
    print("  CEO 보드팩(초안):", a2["path"])
    mani = json.loads(con.execute("SELECT manifest FROM artifact_register WHERE artifact_id=?", (a2["artifact_id"],)).fetchone()[0])
    print(f"  매니페스트: ref_versions={mani['ref_versions']}, models={len(mani['model_versions'])}개, ledger_decisions={mani['ledger_decisions']}")
    print("  artifact_register:")
    for r in con.execute("SELECT period,audience,atype,scope,version,status FROM artifact_register ORDER BY artifact_id"):
        print("   ", r)

    # ----- 16) 외부 검토 반영 개선 (Splink·promptfoo/DeepEval·FP&A-agent 벤치마크) -----
    _line("16) 외부 검토 반영 — ER u-추정 / eval baseline·비용 회귀 / 예측 정확도")
    # (A) ER u 직접추정: 사업자번호 대부분 고유한 스케일 vendor에서 u 작음=강한 증거
    big = [{"name": f"법인{i}", "biz_no": f"{100+i}-00-0000{i%10}"} for i in range(40)] + \
          [{"name": "쿠팡로지스틱스", "biz_no": "123-45-67890"} for _ in range(3)]
    cal = calibrate_u_from_data(big, fields=("biz_no",))
    print(f"  (A) ER u 추정(n={cal['n']}): biz_no u={cal['u_probabilities']['biz_no']} "
          f"→ 일치 가중치 log2(0.99/u)≈{fs_match_weight(0.99, cal['u_probabilities']['biz_no'], True):.1f} bits (강한 증거)")
    print(f"      λ=0.01 사전가중치 M_prior={fs_prior_weight(0.01):.1f} bits  | (Splink: u 직접추정·m은 EM)")
    # (B) eval baseline·비용 회귀
    eval_router(con)  # baseline 적재
    reg = eval_regression_vs_baseline(con, "router_match", value=0.80, le=False)   # 0.80은 직전 1.0 대비 악화
    print(f"  (B) eval baseline 회귀: router 0.80 vs base {reg['baseline']} → regressed={reg['regressed']}")
    print(f"      비용 회귀(예산 $0.01): {eval_cost_regression(con, 0.01)}")
    # (C) 예측 정확도/편향 추적
    for p, f, a in [("2026-03", 168e6, 170e6), ("2026-04", 168e6, 175e6), ("2026-05", 168e6, 185e6)]:
        record_forecast_actual(con, "HQ_lease_pangyo", p, f, a)
    fa = forecast_accuracy(con, "HQ_lease_pangyo", mape_threshold=0.05)
    print(f"  (C) 예측 정확도(HQ_lease): MAPE={fa['mape']:.1%}, bias={fa['bias']:+.1%} → {fa['status']} (실측 대비 과소예측 누적)")

    # ----- 17) 구현 보강·정밀화 (이전에 표준·방법론 기반으로 표기했던 항목을 실제 구현) -----
    _line("17) 구현 보강·정밀화 — ER EM학습 / 하이브리드 검증 / 외부 어댑터 / ABC / 발송 sink / IBR 기본값")
    # B6: ER m을 EM으로 학습(스케일에서 자동 적용) — resolve_entities가 사용
    em = train_er_em(big, "biz_no")
    print(f"  B6 ER EM 학습: m={em['m']}, u={em['u']}, λ={em['lambda']} (n_pairs={em['n_pairs']}) → 가중치 {fs_match_weight(em['m'], em['u'], True):.1f} bits")
    er_big = resolve_entities(con, [{"id": 900 + i, "name": f"법인{i}", "biz_no": f"{700+i}-00-0000{i%10}"} for i in range(31)] +
                              [{"id": 999, "name": "쿠팡로지스틱스", "biz_no": "700-00-00000"}])
    print(f"     스케일 클러스터링(n=32, EM 가중치 사용): 군집 수 {len(er_big['clusters'])}")
    # A2: 하이브리드 클레임 검증(수치+어휘)
    ev = "본사 임차료가 예산 대비 6,600만원 감소했고 주 요인은 재협상이다"
    print(f"  A2 검증(수치 일치): {verify_claim(ev, '임차료 6600만원 감소')['reason']}")
    print(f"  A2 검증(어휘, 정상): {verify_claim(ev, '재협상이 주 요인')['reason']}")
    print(f"  A2 검증(환각 차단): {verify_claim(ev, '신규 차량 도입이 원인이다')['reason']}")
    # B3/A4: 외부 데이터 어댑터(오프라인 file_fetcher로 end-to-end 적재)
    snap = "/tmp/reb_snapshot.json"
    json.dump([{"region": "성남분당", "property_type": "office", "grade": 4, "period": "2026Q2",
                "rent_per_sqm": 20400, "rent_index": 103.1, "vacancy": 3.5}], open(snap, "w", encoding="utf-8"))
    ing = fetch_and_ingest(con, "regional_rent_benchmark", ["region", "property_type", "grade", "period"],
                           "REB_office_rent", "한국부동산원", "data.go.kr", "KOGL-1", "2026Q2", lambda: file_fetcher(snap))
    print(f"  B3 외부 어댑터(오프라인 file_fetcher): {ing['ref_set']} v{ing['version']} 적재 {ing['rows']}행 (api_fetcher는 동일 시그니처)")
    # B7: ABC 일반 배부
    abc = allocate_cost_abc(1_200_000_000, {"DC_서울": 5000, "DC_부산": 3000, "DC_대구": 2000})
    print(f"  B7 ABC 배부(공통비 ₩1.2B, 면적 driver): {abc}")
    # B2: 발송 sink(no-op 아님 — sent_log 실제 기록)
    con.execute("""INSERT INTO task_card(analysis_id,task_type,payload,risk_tier,status,requester,approver,created_at)
        VALUES(NULL,'send_request','{}','low','pending_approval','analyst',NULL,'2026-06-14')""")
    tid = con.execute("SELECT MAX(id) FROM task_card").fetchone()[0]
    approve_task(con, tid, approver="controller")
    sent = process_outbox(con)
    nlog = con.execute("SELECT COUNT(*) FROM sent_log").fetchone()[0]
    print(f"  B2 발송 sink: outbox {len(sent)}건 실행 → sent_log {nlog}행 기록(멱등). 외부 sender 교체 지점.")
    # C1: 완전 IBR 매트릭스 기본값
    seed_full_ibr_matrix(con)
    nibr = con.execute("SELECT COUNT(*) FROM ibr_matrix WHERE valid_to IS NULL").fetchone()[0]
    print(f"  C1 IBR 매트릭스 기본값: {nibr}행(KRW 기간×담보, 플레이스홀더 아님) | C2 hurdle={DEFAULT_HURDLE:.0%} | C6 house_style={HOUSE_STYLE['title_rule']} | C7 보존(boardpack)={RETENTION_POLICY['boardpack']}")

    # ----- 18) 보고 덱/문서 콘텐츠 스펙 (B4: grounded·house_style; 렌더는 BIGS/academic-slide) -----
    _line("18) 보고 덱/문서 콘텐츠 스펙 (B4) — 결론 제목·exhibit·근거 각주·그라운딩")
    deck = build_board_deck_spec(con, "HQ_lease_pangyo", "2026-06")
    print(f"  덱 스펙: 슬라이드 {len(deck['slides'])}매, grounded={deck['grounded']}, 렌더러={deck['renderer']}")
    for s in deck["slides"]:
        print(f"   [{s['layout']}] {s['action_title']}  (출처 {s['source']})")
    rep = build_report_spec(con, "HQ_lease_pangyo", "2026-06")
    print(f"  문서 스펙({rep['format']}): 섹션 {len(rep['sections'])}, grounded={rep['grounded']}, 보존 {rep['retention']}")

if __name__ == "__main__":
    main()
